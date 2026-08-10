"""Offline compilation of configured standards into a candidate Knowledge Pack."""

from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from standards_driven_sdtm_adam.standards import StandardManifest, StandardsRegistry


KNOWLEDGE_PACK_SCHEMA_VERSION = 1
KNOWLEDGE_COMPILER_VERSION = "m3r-1"
SEMANTIC_RECONSTRUCTION_SCHEMA_VERSION = "m3r-phase-3c-r-v1"
V1_DATASETS = ("ADSL", "ADAE", "ADLB", "ADTTE")
CONFLICT_LINEAGE_SCOPES = (
    "ADAE",
    "ADSL",
    "ADSL.RACE",
    "ADSL.SITEID",
    "ADSL.SUBJID",
    "ADSL.TRTEDT",
    "ADSL.TRTSDT",
    "ADTTE",
    "ADLB",
)
PRECEDENCE = (
    "STANDARD_REQUIRED",
    "STANDARD_GUIDED",
    "STUDY_SPECIFIC",
    "USER_DEFINED",
    "DATA_ENGINEERING",
    "DATA_DRIVEN",
)
STUDY_DECISION_TOPICS = (
    "treatment windows",
    "event terms",
    "origins",
    "baseline windows",
    "imputation policies",
)
M3C_R1_SCHEMA_VERSION = "m3c-r1-candidate-audit-v1"
M3C_R2_SCHEMA_VERSION = "m3c-r2-candidate-audit-v1"
M3C_R2A_SCHEMA_VERSION = "m3c-r2a-candidate-audit-v1"
M3C_R2C_GOLD_RULE_PROMOTION_SCHEMA_VERSION = "m3c-r2c-gold-rule-promotion-v1"
GOLD_RULE_PACK_SCHEMA_VERSION = "gold-rule-pack-v1"
NORMATIVE_CUE_TOKENS = (" must ", " shall ", " required", "should ", " may ", "one record per")
MONTH_DATE_PATTERN = re.compile(
    r"(?<![A-Za-z])may(?![A-Za-z])\s+(?:[0-3]?\d(?:st|nd|rd|th)?(?:,\s*|\s+)\d{4}|\d{4})",
    re.IGNORECASE,
)
SOURCE_SECTION_CLASSIFICATIONS = (
    "OPERATIVE_RULE",
    "DEFINITION",
    "CONTROLLED_TERMINOLOGY",
    "EXAMPLE",
    "GLOSSARY",
    "REVISION_HISTORY",
    "FRONT_MATTER",
    "SUBMISSION_CONTEXT",
    "UNRESOLVED",
)
NORMATIVE_SOURCE_SECTIONS = (
    "OPERATIVE_RULE",
    "DEFINITION",
    "CONTROLLED_TERMINOLOGY",
)
NON_OPERATIVE_SOURCE_SECTIONS = (
    "EXAMPLE",
    "GLOSSARY",
    "REVISION_HISTORY",
    "FRONT_MATTER",
    "SUBMISSION_CONTEXT",
    "UNRESOLVED",
)


@dataclass(frozen=True)
class KnowledgePackManifest:
    """Identity and freshness metadata for a compiled Knowledge Pack."""

    schema_version: int
    compiler_version: str
    pack_version: str
    source_registry_root: str | None
    source_manifest_fingerprints: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class CompiledRule:
    """One compiled rule or unresolved content item from a local standard."""

    rule_id: str
    standard_id: str
    standard_name: str
    standard_version: str | None
    classification: str | None
    scope: dict[str, Any]
    applicability_conditions: tuple[str, ...]
    required_behavior: str | None
    permitted_variations: tuple[str, ...]
    prohibited_behavior: tuple[str, ...]
    required_inputs: tuple[str, ...]
    expected_outputs: tuple[str, ...]
    grain: str | None
    keys: tuple[str, ...]
    validation_requirements: tuple[str, ...]
    evidence_locator: dict[str, Any]
    local_relative_source_path: str | None
    official_url: str | None
    source_hash: str | None
    extraction_status: str
    review_status: str
    classification_status: str = "UNRESOLVED"
    normative_strength: str | None = None
    source_section_classification: str | None = None
    exact_source_excerpt: str | None = None
    normalized_atomic_requirement: str | None = None
    rule_type: str | None = None
    dataset_scope: tuple[str, ...] = ()
    structure_scope: tuple[str, ...] = ()
    variable_scope: tuple[str, ...] = ()
    validation_requirement: str | None = None
    semantic_reconstruction_status: str | None = None
    excluded_reason: str | None = None
    conformance_metadata: dict[str, Any] | None = None


@dataclass(frozen=True)
class CompiledKnowledgePack:
    """Compiled reusable standards knowledge."""

    manifest: KnowledgePackManifest
    rules: tuple[CompiledRule, ...]
BATCH_001_ADJUDICATION_FINDINGS = (
    ("ADJ2-0001", "adam-model:07f6b4e8e1592086", "REJECT", None),
    ("ADJ2-0002", "adam-model:9b552da37cd3119e", "RECONSTRUCT", "ADSL one-record-per-subject structure rule"),
    ("ADJ2-0003", "adam-msg:803ef9cbfface96e", "REJECT", None),
    ("ADJ2-0004", "adamig:77ea80f356339175", "EXCLUDE_FROM_V1_CORE", None),
    ("ADJ2-0005", "adam-traceability-examples:98eef8b82139df83", "NON_NORMATIVE_REFERENCE_ONLY", None),
    ("ADJ2-0006", "adam-conformance-rules:sheet-Rules Catalogue_row-728", "RECONSTRUCT", "ADSL.USUBJID required rule"),
    ("ADJ2-0007", "adam-conformance-rules:sheet-Rules Catalogue_row-892", "RECONSTRUCT", "ADSL dataset-name/label rule"),
    ("ADJ2-0008", "adam-conformance-rules:sheet-Rules Catalogue_row-896", "RECONSTRUCT", "inverse ADSL label/name rule"),
    ("ADJ2-0009", "adam-occds:04effc20cb27142a", "REJECT", None),
    ("ADJ2-0010", "adam-occds:45f5a09eb0f62b77", "REJECT", None),
    ("ADJ2-0011", "adam-occds:fa5ce4d8b7728ebf", "REJECT", None),
    ("ADJ2-0012", "adam-model:7c9bd5408ac727f6", "REJECT_CURRENT_AND_SPLIT_INTO_ATOMIC_NAMING_RULES", None),
)


class CandidateExtractionError(RuntimeError):
    """Raised when local candidate extraction fails."""


class KnowledgePackCompiler:
    """Compile complete local standards documents into candidate metadata."""

    def __init__(self, registry: StandardsRegistry) -> None:
        self.registry = registry
        self.unavailable_or_mismatched: list[dict[str, Any]] = []
        self.normalization_audit: dict[str, Any] = {
            "raw_candidate_count": 0,
            "retained_candidate_count": 0,
            "removed_fragment_count": 0,
            "merged_duplicate_count": 0,
            "removed_fragments": [],
            "merged_duplicates": [],
        }

    @classmethod
    def from_registry_dir(cls, registry_dir: str | Path) -> "KnowledgePackCompiler":
        """Load a standards registry for offline compilation."""

        return cls(StandardsRegistry.load(registry_dir, validate_integrity=False))

    def compile(self, *, pack_version: str = "local-candidate") -> CompiledKnowledgePack:
        """Compile all enabled source standards without promoting rules for runtime."""

        self.unavailable_or_mismatched = []
        self.normalization_audit = {
            "raw_candidate_count": 0,
            "retained_candidate_count": 0,
            "removed_fragment_count": 0,
            "merged_duplicate_count": 0,
            "removed_fragments": [],
            "merged_duplicates": [],
        }
        rules: list[CompiledRule] = []
        fingerprints = []
        extracted_at = _utc_now()
        for manifest in self.registry.enabled():
            integrity = self._integrity(manifest)
            fingerprints.append(_fingerprint(manifest, integrity, extracted_at))
            if integrity["status"] != "OK":
                self.unavailable_or_mismatched.append(_problem(manifest, integrity))
                rules.append(_status_rule(manifest, integrity, extracted_at))
                continue

            source_path = self.registry.resolve_local_path(manifest)
            assert source_path is not None
            try:
                candidates = _extract_candidates(manifest, source_path, self.registry.root)
            except (OSError, CandidateExtractionError) as exc:
                integrity = dict(integrity)
                integrity["status"] = "TEXT_EXTRACTION_FAILED"
                integrity["reason"] = str(exc)
                self.unavailable_or_mismatched.append(_problem(manifest, integrity))
                rules.append(_status_rule(manifest, integrity, extracted_at))
                continue

            self.normalization_audit["raw_candidate_count"] += len(candidates)
            normalized = _normalize_candidates(manifest, candidates)
            self.normalization_audit["removed_fragment_count"] += len(normalized["removed_fragments"])
            self.normalization_audit["merged_duplicate_count"] += len(normalized["merged_duplicates"])
            self.normalization_audit["removed_fragments"].extend(normalized["removed_fragments"])
            self.normalization_audit["merged_duplicates"].extend(normalized["merged_duplicates"])
            for index, candidate in enumerate(normalized["retained"], start=1):
                rules.append(_candidate_rule(manifest, source_path, self.registry.root, candidate, index, extracted_at))

        self.normalization_audit["retained_candidate_count"] = sum(
            1 for rule in rules if rule.extraction_status == "CANDIDATE_RULE"
        )
        manifest = KnowledgePackManifest(
            schema_version=KNOWLEDGE_PACK_SCHEMA_VERSION,
            compiler_version=KNOWLEDGE_COMPILER_VERSION,
            pack_version=pack_version,
            source_registry_root=str(self.registry.root),
            source_manifest_fingerprints=tuple(fingerprints),
        )
        return CompiledKnowledgePack(manifest=manifest, rules=tuple(rules))

    def write(self, output_root: str | Path, *, pack_version: str = "local-candidate") -> Path:
        """Write a candidate pack to the canonical directory structure."""

        pack = self.compile(pack_version=pack_version)
        root = Path(output_root)
        for child in (
            "rule_catalog",
            "dataset_structures",
            "variable_requirements",
            "conformance_rules",
            "rule_precedence",
            "coverage_matrix",
            "citation_index",
            "review_queue",
            "gold_rules",
        ):
            (root / child).mkdir(parents=True, exist_ok=True)

        rules = [asdict(rule) for rule in pack.rules]
        summary = _summary(pack, self.unavailable_or_mismatched, self.normalization_audit)
        (root / "manifest.json").write_text(
            json.dumps({"manifest": asdict(pack.manifest), "compilation_summary": summary}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        for old_shard in (root / "rule_catalog").glob("*.json"):
            old_shard.unlink()
        shard_index = _write_rule_shards(root / "rule_catalog", rules)
        (root / "rule_catalog" / "index.json").write_text(
            json.dumps({"shards": shard_index}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "dataset_structures" / "dataset_structures.json").write_text(
            json.dumps({"dataset_structures": _dataset_structures(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "variable_requirements" / "variable_requirements.json").write_text(
            json.dumps({"variable_requirements": _variable_requirements(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "conformance_rules" / "conformance_rules.json").write_text(
            json.dumps({"conformance_rules": _conformance_rules(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "rule_precedence" / "rule_precedence.json").write_text(
            json.dumps({"rule_precedence": _rule_precedence(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "coverage_matrix" / "v1_coverage_matrix.json").write_text(
            json.dumps({"coverage": _coverage_matrix(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "citation_index" / "citation_index.json").write_text(
            json.dumps({"citations": _citation_index(pack.rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        review_queue = _review_queue(pack.rules)
        conflicts = _conflict_review(pack.rules, _rule_precedence(pack.rules)["detected_conflicts"])
        gap_map = _coverage_gap_review_map(_coverage_matrix(pack.rules), review_queue)
        (root / "review_queue" / "v1_core_review_queue.json").write_text(
            json.dumps({"review_items": review_queue}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "review_queue" / "conflict_review.json").write_text(
            json.dumps({"conflict_reviews": conflicts}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "review_queue" / "coverage_gap_map.json").write_text(
            json.dumps({"coverage_gaps": gap_map}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "review_queue" / "review_input_freeze.json").write_text(
            json.dumps(_review_input_freeze(root, pack, shard_index), indent=2, sort_keys=True),
            encoding="utf-8",
        )
        for existing in (root / "gold_rules").iterdir():
            if existing.is_file():
                existing.unlink()
        (root / "compilation_summary.json").write_text(
            json.dumps(summary, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        return root

    def write_semantic_reconstruction(
        self,
        output_root: str | Path,
        *,
        pack_version: str = "m3r-phase-3c-r-semantic-reconstruction-v1",
    ) -> Path:
        """Write a reconstructed candidate pack without replacing earlier review artifacts."""

        pack = self.compile(pack_version=pack_version)
        root = Path(output_root)
        reconstructed_root = root / "v1"
        for child in (
            "rule_catalog",
            "coverage_matrix",
            "citation_index",
            "conformance_rules",
        ):
            (reconstructed_root / child).mkdir(parents=True, exist_ok=True)
        (root / "review_queue").mkdir(parents=True, exist_ok=True)
        (root / "gold_rules").mkdir(parents=True, exist_ok=True)

        reconstructed_rules = _reconstructed_rules(pack.rules)
        reconstructed_rules = _apply_batch_001_artifact_reconstructions(root, pack, reconstructed_rules)
        rules_payload = [_semantic_rule_payload(rule) for rule in reconstructed_rules]
        excluded = _semantic_exclusions(pack.rules)
        summary = _semantic_reconstruction_summary(pack.rules, reconstructed_rules, excluded, self.normalization_audit)
        shard_index = _write_rule_shards(reconstructed_root / "rule_catalog", rules_payload)
        (reconstructed_root / "rule_catalog" / "index.json").write_text(
            json.dumps({"shards": shard_index}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (reconstructed_root / "manifest.json").write_text(
            json.dumps(
                {
                    "manifest": asdict(pack.manifest),
                    "schema_version": SEMANTIC_RECONSTRUCTION_SCHEMA_VERSION,
                    "summary": summary,
                    "supersedes_review_generation": "SUPERSEDED_BY_CANDIDATE_RECONSTRUCTION",
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        (reconstructed_root / "compilation_summary.json").write_text(
            json.dumps(summary, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (reconstructed_root / "citation_index" / "citation_index.json").write_text(
            json.dumps({"citations": _citation_index(reconstructed_rules)}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (reconstructed_root / "coverage_matrix" / "v1_coverage_matrix.json").write_text(
            json.dumps({"coverage": _coverage_matrix(tuple(reconstructed_rules))}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (reconstructed_root / "conformance_rules" / "conformance_rules.json").write_text(
            json.dumps({"conformance_rules": _conformance_rules(tuple(reconstructed_rules))}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (reconstructed_root / "excluded_candidates.json").write_text(
            json.dumps({"excluded_candidates": excluded}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        review_queue = _semantic_review_queue(tuple(reconstructed_rules), limit=180)
        (root / "review_queue" / "reconstructed_v1_core_review_queue.json").write_text(
            json.dumps({"review_items": review_queue}, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        (root / "review_queue" / "batch_001_adjudication_findings.json").write_text(
            json.dumps(
                {
                    "schema_version": "m3r-phase-3c-r-batch-001-findings-v1",
                    "review_generation_status": "SUPERSEDED_BY_CANDIDATE_RECONSTRUCTION",
                    "findings": [
                        {
                            "adjudication_item_id": item_id,
                            "candidate_rule_id": candidate_id,
                            "decision": decision,
                            "reconstruction_instruction": instruction,
                            "promotion_authorized": False,
                        }
                        for item_id, candidate_id, decision, instruction in BATCH_001_ADJUDICATION_FINDINGS
                    ],
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        (root / "review_queue" / "review_generation_supersession.json").write_text(
            json.dumps(
                {
                    "schema_version": "m3r-phase-3c-r-supersession-v1",
                    "status": "SUPERSEDED_BY_CANDIDATE_RECONSTRUCTION",
                    "superseded_artifacts": (
                        "reviewer_a_results.json",
                        "reviewer_b_results.json",
                        "dual_review_comparison.json",
                        "dual_review_comparison_corrected_v1.json",
                        "dual_review_comparison_corrected_v2.json",
                        "proposed_revised_candidates_v1.json",
                        "proposed_revised_candidates_v2.json",
                        "adjudication_batches/",
                        "adjudication_batches_v2/",
                    ),
                    "preservation_policy": "Existing artifacts are preserved unchanged and must not be used for Gold Rule approval.",
                    "created_at": _utc_now(),
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        (root / "review_queue" / "reconstructed_review_input_freeze.json").write_text(
            json.dumps(_semantic_review_input_freeze(root, reconstructed_root, pack, shard_index), indent=2, sort_keys=True),
            encoding="utf-8",
        )
        return reconstructed_root

    def write_m3c_r1_audit(
        self,
        output_root: str | Path,
        *,
        audit_version: str = "m3c-r1",
        schema_version: str = M3C_R1_SCHEMA_VERSION,
        review_queue_filename: str = "m3c_r1_adsl_structure_grain_key_candidates.json",
        pack_version: str = "m3c-r1-candidate-reconciliation-v1",
    ) -> Path:
        """Write versioned M3C reconciliation and review-preparation artifacts."""

        pack = self.compile(pack_version=pack_version)
        root = Path(output_root)
        audit_root = root / "reconstructed_candidate_pack" / audit_version
        review_root = root / "review_queue"
        audit_root.mkdir(parents=True, exist_ok=True)
        review_root.mkdir(parents=True, exist_ok=True)
        (root / "gold_rules").mkdir(parents=True, exist_ok=True)

        base_reconstructed = _reconstructed_rules(pack.rules)
        reconstructed = _apply_batch_001_artifact_reconstructions(root, pack, base_reconstructed)
        excluded = _semantic_exclusions(pack.rules)
        citation_unresolved = [rule for rule in reconstructed if not _citation_provenance_complete(rule)]
        citation_resolved = [rule for rule in reconstructed if _citation_provenance_complete(rule)]
        review_candidates = _adsl_structure_grain_key_candidates(tuple(citation_resolved), limit=10)

        artifacts: dict[Path, dict[str, Any]] = {
            audit_root / "candidate_count_reconciliation.json": _m3c_r1_reconciliation(
                pack.rules,
                base_reconstructed,
                reconstructed,
                excluded,
                citation_unresolved,
                self.normalization_audit,
                schema_version=schema_version,
            ),
            audit_root / "citation_unresolved_candidates.json": {
                "schema_version": schema_version,
                "status": "CITATION_UNRESOLVED",
                "candidate_count": len(citation_unresolved),
                "review_queue_admitted_count": 0,
                "runtime_pack_admitted_count": 0,
                "candidates": [_citation_unresolved_record(rule) for rule in citation_unresolved],
            },
            audit_root / "source_section_spot_checks.json": _source_section_spot_checks(
                pack.rules,
                self.normalization_audit,
                sample_size=20,
                schema_version=schema_version,
            ),
            review_root / review_queue_filename: {
                "schema_version": schema_version,
                "selection_basis": "ADSL structure, grain, and key coverage with complete citation provenance; duplicates by official conformance rule identifier are collapsed for review preparation.",
                "candidate_count": len(review_candidates),
                "review_status": "CANDIDATE",
                "promotion_authorized": False,
                "candidates": review_candidates,
            },
        }
        file_hashes = []
        for path, payload in artifacts.items():
            text = json.dumps(payload, indent=2, sort_keys=True)
            path.write_text(text, encoding="utf-8")
            file_hashes.append(
                {
                    "file_path": path.as_posix(),
                    "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                }
            )
        aggregate_hash = hashlib.sha256(
            json.dumps(file_hashes, sort_keys=True).encode("utf-8")
        ).hexdigest()
        manifest_path = audit_root / "artifact_manifest.json"
        manifest_path.write_text(
            json.dumps(
                {
                    "schema_version": schema_version,
                    "artifact_hash": aggregate_hash,
                    "artifact_hash_scope": "Ordered file paths and SHA256 values listed in files.",
                    "files": file_hashes,
                    "candidate_pack_count": len(reconstructed),
                    "citation_resolved_candidate_count": len(citation_resolved),
                    "citation_unresolved_candidate_count": len(citation_unresolved),
                    "review_queue_candidate_count": len(review_candidates),
                    "runtime_pack_admitted_count": 0,
                    "review_status": "CANDIDATE",
                    "promotion_authorized": False,
                    "created_at": _utc_now(),
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        return manifest_path

    def write_m3c_r2a_audit(self, output_root: str | Path) -> Path:
        """Write the scoped M3C-R2A ADSL revision artifacts without promotion."""

        root = Path(output_root)
        baseline_manifest_path = root / "reconstructed_candidate_pack" / "m3c-r1" / "artifact_manifest.json"
        baseline_reconciliation_path = (
            root / "reconstructed_candidate_pack" / "m3c-r1" / "candidate_count_reconciliation.json"
        )
        baseline_review_path = root / "review_queue" / "m3c_r1_adsl_structure_grain_key_candidates.json"
        baseline_manifest = _read_json(baseline_manifest_path)
        baseline_reconciliation = _read_json(baseline_reconciliation_path)
        baseline_review = _read_json(baseline_review_path)

        parent_candidates = list(baseline_review.get("candidates") or [])
        revisions = _m3c_r2a_revised_adsl_candidates(parent_candidates, self.registry)
        key_search = _m3c_r2a_adsl_key_evidence_search(parent_candidates, self.registry)
        key_candidate = key_search.get("candidate")
        active_total = int(baseline_manifest["candidate_pack_count"]) + (1 if key_candidate else 0)
        resolved_total = int(baseline_manifest["citation_resolved_candidate_count"]) + (
            1 if key_candidate and key_candidate.get("citation_provenance", {}).get("citation_status") == "RESOLVED" else 0
        )
        unresolved_total = int(baseline_manifest["citation_unresolved_candidate_count"]) + (
            1 if key_candidate and key_candidate.get("citation_provenance", {}).get("citation_status") != "RESOLVED" else 0
        )

        audit_root = root / "reconstructed_candidate_pack" / "m3c-r2a"
        review_root = root / "review_queue"
        audit_root.mkdir(parents=True, exist_ok=True)
        review_root.mkdir(parents=True, exist_ok=True)
        artifacts: dict[Path, dict[str, Any]] = {
            audit_root / "candidate_count_reconciliation.json": _m3c_r2a_reconciliation(
                baseline_manifest,
                baseline_reconciliation,
                revisions,
                key_search,
                active_total,
                resolved_total,
                unresolved_total,
            ),
            audit_root / "source_section_spot_checks.json": _m3c_r2a_source_section_spot_checks(revisions),
            audit_root / "adsl_key_evidence_search.json": key_search,
            audit_root / "revised_adsl_candidates.json": {
                "schema_version": M3C_R2A_SCHEMA_VERSION,
                "status": "REVISED_PENDING_APPROVAL",
                "candidate_count": len(revisions),
                "promotion_authorized": False,
                "runtime_pack_admitted_count": 0,
                "candidates": revisions,
            },
            review_root / "m3c_r2a_adsl_structure_grain_key_candidates.json": {
                "schema_version": M3C_R2A_SCHEMA_VERSION,
                "selection_basis": (
                    "M3C-R2A one-for-one ADSL structure/grain review revisions plus any independently "
                    "discovered formal ADSL key candidate; parent M3C-R1 candidates are superseded only "
                    "for this active review scope."
                ),
                "candidate_count": len(revisions) + (1 if key_candidate else 0),
                "review_status": "REVISED_PENDING_APPROVAL",
                "promotion_authorized": False,
                "runtime_pack_admitted_count": 0,
                "key_evidence_status": key_search["key_evidence_status"],
                "candidates": revisions + ([key_candidate] if key_candidate else []),
            },
            review_root / "m3c_r2a_supersession_manifest.json": _m3c_r2a_supersession_manifest(revisions),
        }

        file_hashes = []
        for path, payload in artifacts.items():
            path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
            file_hashes.append({"file_path": path.as_posix(), "sha256": hashlib.sha256(path.read_bytes()).hexdigest()})
        aggregate_hash = hashlib.sha256(json.dumps(file_hashes, sort_keys=True).encode("utf-8")).hexdigest()
        manifest_path = audit_root / "artifact_manifest.json"
        manifest_path.write_text(
            json.dumps(
                {
                    "schema_version": M3C_R2A_SCHEMA_VERSION,
                    "artifact_hash": aggregate_hash,
                    "artifact_hash_scope": "Ordered file paths and SHA256 values listed in files.",
                    "baseline": {
                        "version": "m3c-r1",
                        "confirmed": True,
                        "artifact_hash": baseline_manifest["artifact_hash"],
                        "candidate_pack_count": baseline_manifest["candidate_pack_count"],
                        "citation_resolved_candidate_count": baseline_manifest["citation_resolved_candidate_count"],
                        "citation_unresolved_candidate_count": baseline_manifest["citation_unresolved_candidate_count"],
                    },
                    "files": file_hashes,
                    "candidate_pack_count": active_total,
                    "citation_resolved_candidate_count": resolved_total,
                    "citation_unresolved_candidate_count": unresolved_total,
                    "adsl_parent_candidates_considered": len(parent_candidates),
                    "adsl_revised_candidates_created": len(revisions),
                    "parent_records_deleted": 0,
                    "gold_rules_written": 0,
                    "runtime_pack_admitted_count": 0,
                    "review_status": "REVISED_PENDING_APPROVAL",
                    "promotion_authorized": False,
                    "created_at": _utc_now(),
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        return manifest_path

    def _integrity(self, manifest: StandardManifest) -> dict[str, Any]:
        if manifest.local_path is None:
            if manifest.local_root is not None:
                missing = self.registry.missing_package_members(manifest)
                return {
                    "status": "PACKAGE_REFERENCE_NOT_PARSED" if not missing else "PACKAGE_MEMBER_MISSING",
                    "reason": "Package-only reference has no single authoritative local document path for candidate parsing."
                    if not missing
                    else f"Missing package members: {', '.join(missing)}",
                    "sha256_status": "NOT_APPLICABLE",
                }
            return {"status": "SOURCE_UNAVAILABLE", "reason": "No local path configured.", "sha256_status": "NOT_APPLICABLE"}
        source_path = self.registry.resolve_local_path(manifest)
        if source_path is None or not source_path.exists():
            return {"status": "SOURCE_UNAVAILABLE", "reason": "Configured local path is missing.", "sha256_status": "MISSING"}
        sha_status = self.registry.sha256_status(manifest)
        if sha_status not in {"PRESENT", "NOT_APPLICABLE"}:
            return {"status": "DOCUMENT_IDENTITY_MISMATCH", "reason": f"SHA256 status is {sha_status}.", "sha256_status": sha_status}
        if manifest.sha256_status == "MISMATCH" or manifest.verification_status == "MISMATCH":
            return {"status": "DOCUMENT_IDENTITY_MISMATCH", "reason": "Manifest records an identity mismatch.", "sha256_status": sha_status}
        return {"status": "OK", "reason": None, "sha256_status": sha_status}


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def promote_m3c_r2c_gold_rules(knowledge_root: str | Path) -> Path:
    """Promote approved M3C-R2C ADSL candidates into citation-only Gold Rules."""

    root = Path(knowledge_root)
    r2a_manifest = _read_json(root / "reconstructed_candidate_pack" / "m3c-r2a" / "artifact_manifest.json")
    r2b_manifest = _read_json(root / "reconstructed_candidate_pack" / "m3c-r2b" / "artifact_manifest.json")
    r2a_review = _read_json(root / "review_queue" / "m3c_r2a_adsl_structure_grain_key_candidates.json")
    r2b_proposal = _read_json(root / "review_queue" / "m3c_r2b_adsl_approval_proposal.json")
    r2c_approval = _read_json(root / "review_queue" / "m3c_r2c_project_owner_approval.json")

    _validate_m3c_r2c_promotion_inputs(r2a_manifest, r2b_manifest, r2b_proposal, r2c_approval)
    approved_by_id = {
        str(item["candidate_id"]): item
        for item in r2c_approval.get("approved_candidates") or []
        if item.get("decision") == "APPROVED"
    }
    r2a_by_id = {str(item["candidate_id"]): item for item in r2a_review.get("candidates") or []}
    proposals = list(r2b_proposal.get("included_proposals") or [])
    gold_rules = [
        _m3c_r2c_gold_rule(proposal, r2a_by_id[str(proposal["candidate_id"])], approved_by_id[str(proposal["candidate_id"])])
        for proposal in proposals
    ]

    gold_root = root / "gold_rules"
    promotion_root = root / "reconstructed_candidate_pack" / "m3c-r2c-promotion"
    gold_root.mkdir(parents=True, exist_ok=True)
    promotion_root.mkdir(parents=True, exist_ok=True)

    gold_rules_path = gold_root / "adsl_m3c_r2c_gold_rules.json"
    gold_payload = {
        "schema_version": GOLD_RULE_PACK_SCHEMA_VERSION,
        "promotion_schema_version": M3C_R2C_GOLD_RULE_PROMOTION_SCHEMA_VERSION,
        "source_approval_artifact": "knowledge/review_queue/m3c_r2c_project_owner_approval.json",
        "rule_count": len(gold_rules),
        "runtime_pack_admitted": False,
        "adsl_formal_key_status": "KEY_EVIDENCE_UNRESOLVED",
        "gold_rules": gold_rules,
    }
    gold_rules_path.write_text(json.dumps(gold_payload, indent=2, sort_keys=True), encoding="utf-8")

    files = [
        {
            "file_path": _knowledge_relative_path(gold_rules_path, root),
            "sha256": hashlib.sha256(gold_rules_path.read_bytes()).hexdigest(),
        }
    ]
    artifact_hash = hashlib.sha256(json.dumps(files, sort_keys=True).encode("utf-8")).hexdigest()
    manifest_path = promotion_root / "artifact_manifest.json"
    manifest_payload = {
        "schema_version": M3C_R2C_GOLD_RULE_PROMOTION_SCHEMA_VERSION,
        "status": "PROMOTED_GOLD_RULES",
        "created_at": _utc_now(),
        "basis_m3c_r2a_aggregate_hash": r2a_manifest["artifact_hash"],
        "basis_m3c_r2b_aggregate_hash": r2b_manifest["artifact_hash"],
        "basis_m3c_r2c_approval_hash": hashlib.sha256(
            (root / "review_queue" / "m3c_r2c_project_owner_approval.json").read_bytes()
        ).hexdigest(),
        "artifact_hash": artifact_hash,
        "artifact_hash_scope": "Ordered file paths and SHA256 values listed in files.",
        "files": files,
        "approved_candidate_count": len(approved_by_id),
        "gold_rules_written": len(gold_rules),
        "adsl_formal_key_status": "KEY_EVIDENCE_UNRESOLVED",
        "coverage_gates": {
            "ADSL_KEY_EVIDENCE_UNRESOLVED": "BLOCKED",
            "EXECUTABLE_SPECIFICATION_BLOCKED": "BLOCKED",
            "V1_READY_BLOCKED": "BLOCKED",
        },
        "operator_binding_performed": False,
        "specification_generated": False,
        "transfer_run": False,
        "formal_report_generated": False,
        "runtime_pack_admitted": False,
        "historical_candidate_artifacts_modified": False,
    }
    manifest_path.write_text(json.dumps(manifest_payload, indent=2, sort_keys=True), encoding="utf-8")
    return manifest_path


def _validate_m3c_r2c_promotion_inputs(
    r2a_manifest: dict[str, Any],
    r2b_manifest: dict[str, Any],
    r2b_proposal: dict[str, Any],
    r2c_approval: dict[str, Any],
) -> None:
    if r2a_manifest.get("artifact_hash") != "ac3155e5ed1eedbf13077bd64d56d7caa236145ac71678e0c56df760a94c8a3e":
        raise ValueError("M3C-R2A artifact hash does not match the approved promotion basis.")
    if r2b_manifest.get("artifact_hash") != "afa2893b877aa979858ed8d342ec8d001547a329121ae56ecee4e7ab17590f23":
        raise ValueError("M3C-R2B artifact hash does not match the approved promotion basis.")
    if r2c_approval.get("adsl_formal_key_decision", {}).get("status") != "KEY_EVIDENCE_UNRESOLVED":
        raise ValueError("ADSL formal key must remain KEY_EVIDENCE_UNRESOLVED.")

    proposals = list(r2b_proposal.get("included_proposals") or [])
    approvals = list(r2c_approval.get("approved_candidates") or [])
    approved_ids = {str(item.get("candidate_id")) for item in approvals if item.get("decision") == "APPROVED"}
    proposal_ids = {str(item.get("candidate_id")) for item in proposals}
    if len(proposals) != 5 or len(approved_ids) != 5 or proposal_ids != approved_ids:
        raise ValueError("M3C-R2C approval must map one-to-one to the five M3C-R2B included proposals.")

    for proposal in proposals:
        if proposal.get("classification") != "STANDARD_REQUIRED":
            raise ValueError(f"Gold Rule candidate is not STANDARD_REQUIRED: {proposal.get('candidate_id')}")
        citation = proposal.get("citation_summary") or {}
        required_citation_fields = (
            "citation_status",
            "standard_title",
            "standard_version",
            "local_relative_path",
            "official_url",
            "locator",
            "source_sha256",
        )
        if any(citation.get(field) in (None, "", {}) for field in required_citation_fields):
            raise ValueError(f"Gold Rule candidate has incomplete citation: {proposal.get('candidate_id')}")
        if citation.get("citation_status") != "RESOLVED":
            raise ValueError(f"Gold Rule candidate citation is not resolved: {proposal.get('candidate_id')}")
        if Path(str(citation.get("local_relative_path"))).is_absolute():
            raise ValueError(f"Gold Rule citation path must be repository-relative: {proposal.get('candidate_id')}")
        for prohibited in ("operator", "operator_eligibility", "operator_parameters"):
            if prohibited in proposal:
                raise ValueError(f"Gold Rule candidate contains prohibited field {prohibited}: {proposal.get('candidate_id')}")


def _m3c_r2c_gold_rule(
    proposal: dict[str, Any],
    candidate: dict[str, Any],
    approval: dict[str, Any],
) -> dict[str, Any]:
    candidate_id = str(proposal["candidate_id"])
    citation = dict(proposal["citation_summary"])
    return {
        "gold_rule_id": f"gold:{_hash_text(candidate_id)}",
        "candidate_id": candidate_id,
        "classification": proposal["classification"],
        "behavior_type": candidate.get("behavior_type") or "REQUIRED",
        "atomic_rule_text": proposal["atomic_rule_text"],
        "dataset_scope": list(candidate.get("dataset_scope") or []),
        "structure_scope": candidate.get("structure_scope"),
        "variable_scope": list(candidate.get("variable_scope") or []),
        "applicability_conditions": list(candidate.get("applicability_conditions") or []),
        "citation": {
            "citation_status": citation["citation_status"],
            "standard_title": citation["standard_title"],
            "standard_version": citation["standard_version"],
            "local_relative_path": citation["local_relative_path"],
            "official_url": citation["official_url"],
            "locator": citation["locator"],
            "source_sha256": citation["source_sha256"],
        },
        "lineage": {
            "promoted_from": "m3c-r2a",
            "approval_proposal": "knowledge/review_queue/m3c_r2b_adsl_approval_proposal.json",
            "project_owner_approval": "knowledge/review_queue/m3c_r2c_project_owner_approval.json",
            "parent_candidate_id": candidate.get("parent_candidate_id"),
        },
        "approval": {
            "reviewer_role": "PROJECT_OWNER",
            "approval_source": "EXPLICIT_USER_INSTRUCTION",
            "decision": approval["decision"],
            "approval_basis": approval.get("approval_basis"),
        },
    }


def _knowledge_relative_path(path: Path, knowledge_root: Path) -> str:
    try:
        return f"knowledge/{path.resolve().relative_to(knowledge_root.resolve()).as_posix()}"
    except ValueError:
        return path.as_posix()


def _m3c_r2a_revised_adsl_candidates(
    parents: list[dict[str, Any]],
    registry: StandardsRegistry,
) -> list[dict[str, Any]]:
    revision_specs = {
        "adam-model:9b552da37cd3119e:reconstructed-adsl-one-record-per-subject": {
            "suffix": "m3c-r2a-revised-adsl-one-record-per-subject",
            "requirement": "ADSL must contain exactly one record per subject.",
            "revision_reason": "clarify atomic subject-level grain requirement; do not infer formal dataset key.",
        },
        "adam-conformance-rules:sheet-Rules Catalogue_row-728:reconstructed-adsl-usubjid-required": {
            "suffix": "m3c-r2a-revised-adsl-usubjid-populated",
            "requirement": "When USUBJID is present in ADSL, every ADSL record must have a populated USUBJID value.",
            "revision_reason": (
                "preserve Rule 256 condition and required populated-USUBJID behavior without broadening into "
                "formal key evidence."
            ),
        },
        "adam-conformance-rules:sheet-Rules Catalogue_row-892:reconstructed-adsl-name-label": {
            "suffix": "m3c-r2a-revised-adsl-name-label",
            "requirement": 'A dataset named ADSL must have the dataset label "Subject-Level Analysis Dataset".',
            "revision_reason": "preserve Rule 320 name-to-label rule and remove unsupported applicability condition.",
        },
        "adam-conformance-rules:sheet-Rules Catalogue_row-896:reconstructed-adsl-label-name": {
            "suffix": "m3c-r2a-revised-adsl-label-name",
            "requirement": 'A dataset with label "Subject-Level Analysis Dataset" must be named ADSL.',
            "revision_reason": "preserve Rule 321 label-to-name rule and remove unsupported applicability condition.",
        },
        "adam-conformance-rules:sheet-Rules Catalogue_row-199": {
            "suffix": "m3c-r2a-revised-adsl-usubjid-uniqueness",
            "requirement": "ADSL must not contain more than one record for the same USUBJID.",
            "revision_reason": "normalize negative duplicate-record rule into atomic uniqueness behavior without inferring full dataset key.",
        },
    }
    by_id = {str(parent.get("candidate_id")): parent for parent in parents}
    revisions: list[dict[str, Any]] = []
    for parent_id, spec in revision_specs.items():
        parent = by_id[parent_id]
        revision_id = f"{parent_id}:{spec['suffix']}"
        removed_conditions = []
        if parent_id.endswith(("reconstructed-adsl-name-label", "reconstructed-adsl-label-name")):
            removed_conditions = list(parent.get("applicability_conditions") or [])
        citation = _canonical_m3c_r2a_citation(dict(parent.get("citation_provenance") or {}), registry)
        verification = _verify_m3c_r2a_citation(citation, registry)
        citation["citation_status"] = "RESOLVED" if verification["verification_status"] == "VERIFIED" else "CITATION_UNRESOLVED"
        revision = {
            "candidate_id": revision_id,
            "parent_candidate_id": parent_id,
            "exact_source_excerpt": parent.get("exact_source_excerpt"),
            "normalized_atomic_requirement": spec["requirement"],
            "classification": "STANDARD_REQUIRED",
            "behavior_type": "REQUIRED",
            "normative_strength": "REQUIRED",
            "applicability_conditions": [],
            "dataset_scope": ["ADSL"],
            "structure_scope": "SUBJECT_LEVEL_ANALYSIS_DATASET",
            "variable_scope": parent.get("variable_scope") or [],
            "citation_provenance": citation,
            "citation_verification": verification,
            "source_sha256": citation.get("source_sha256"),
            "review_status": "CANDIDATE",
            "semantic_review_status": "REVISED_PENDING_APPROVAL",
            "revision_reason": spec["revision_reason"],
            "promotion_authorized": False,
            "runtime_pack_admitted": False,
            "lineage": {
                "parent_candidate_id": parent_id,
                "revision_candidate_id": revision_id,
                "lineage_status": "SUPERSEDES_PARENT_FOR_M3C_R2A_REVIEW",
                "parent_record_preserved": True,
                "parent_active_in_m3c_r2a": False,
                "revision_active_in_m3c_r2a": True,
                "promotion_authorized": False,
                "runtime_pack_admitted": False,
            },
        }
        if removed_conditions:
            revision["removed_applicability_conditions"] = removed_conditions
            revision["removal_reason"] = "no independent citation support attached to this candidate"
        revisions.append(revision)
    return revisions


def _verify_m3c_r2a_citation(citation: dict[str, Any], registry: StandardsRegistry) -> dict[str, Any]:
    standard_id = str(citation.get("standard_id") or "")
    required = ("standard_title", "standard_version", "local_relative_path", "official_url", "locator", "source_sha256")
    missing = [field for field in required if not citation.get(field)]
    observed: dict[str, Any] = {}
    status = "VERIFIED"
    try:
        manifest = registry.get(standard_id)
        source_path = registry.resolve_local_path(manifest)
        observed.update(
            {
                "standard_title": manifest.title,
                "standard_version": manifest.version,
                "local_relative_path": _relative_source(registry.root, source_path) if source_path else None,
                "official_url": manifest.official_url,
                "source_sha256": registry.calculate_sha256(manifest) if source_path and source_path.exists() else None,
                "registry_sha256": manifest.sha256,
                "sha256_status": registry.sha256_status(manifest),
            }
        )
    except Exception as exc:  # pragma: no cover - defensive artifact metadata
        observed["error"] = str(exc)
        status = "CITATION_UNRESOLVED"
    comparisons = {
        "standard_title_matches_registry": citation.get("standard_title") == observed.get("standard_title"),
        "standard_version_matches_registry": citation.get("standard_version") == observed.get("standard_version"),
        "official_url_matches_registry": citation.get("official_url") == observed.get("official_url"),
        "source_sha256_matches_registry": citation.get("source_sha256") == observed.get("source_sha256"),
        "locator_present": bool(citation.get("locator")),
        "candidate_local_relative_path": citation.get("local_relative_path"),
        "observed_registry_relative_path": observed.get("local_relative_path"),
    }
    if missing or observed.get("sha256_status") != "PRESENT" or not all(
        comparisons[key]
        for key in (
            "standard_title_matches_registry",
            "standard_version_matches_registry",
            "official_url_matches_registry",
            "source_sha256_matches_registry",
            "locator_present",
        )
    ):
        status = "CITATION_UNRESOLVED"
    return {
        "verification_status": status,
        "missing_fields": missing,
        "observed_from_registry_and_source": observed,
        "comparisons": comparisons,
    }


def _canonical_m3c_r2a_citation(citation: dict[str, Any], registry: StandardsRegistry) -> dict[str, Any]:
    standard_id = str(citation.get("standard_id") or "")
    try:
        manifest = registry.get(standard_id)
        source_path = registry.resolve_local_path(manifest)
    except Exception:
        return citation
    if source_path is None:
        return citation
    repo_relative_path = _relative_source(registry.root, source_path)
    if citation.get("local_relative_path") != repo_relative_path:
        citation["local_relative_path"] = repo_relative_path
    return citation


def _m3c_r2a_adsl_key_evidence_search(
    parents: list[dict[str, Any]],
    registry: StandardsRegistry,
) -> dict[str, Any]:
    search_terms = ("ADSL", "key", "keys", "unique", "identifier", "STUDYID", "USUBJID", "Subject-Level Analysis Dataset")
    reviewed_matches = []
    direct_key_candidate = None
    for parent in parents:
        text = " ".join(str(parent.get(field) or "") for field in ("normalized_atomic_requirement", "exact_source_excerpt"))
        if not any(term.lower() in text.lower() for term in search_terms):
            continue
        rejection_reason = _adsl_key_rejection_reason(text)
        match_record = {
            "candidate_id": parent.get("candidate_id"),
            "matched_text": _exact_prefix(text),
            "citation_provenance": _canonical_m3c_r2a_citation(dict(parent.get("citation_provenance") or {}), registry),
            "direct_formal_key_evidence": rejection_reason is None,
            "review_disposition": "REJECTED_AS_INFERENCE_ONLY" if rejection_reason else "DIRECT_FORMAL_KEY_EVIDENCE",
            "reason": rejection_reason,
        }
        reviewed_matches.append(match_record)
        if rejection_reason is None and direct_key_candidate is None:
            citation = _canonical_m3c_r2a_citation(dict(parent.get("citation_provenance") or {}), registry)
            verification = _verify_m3c_r2a_citation(citation, registry)
            citation["citation_status"] = "RESOLVED" if verification["verification_status"] == "VERIFIED" else "CITATION_UNRESOLVED"
            direct_key_candidate = {
                "candidate_id": f"{parent['candidate_id']}:m3c-r2a-independent-adsl-formal-key",
                "parent_candidate_id": parent.get("candidate_id"),
                "normalized_atomic_requirement": "ADSL formal dataset key evidence requires semantic approval before specification use.",
                "classification": "STANDARD_REQUIRED",
                "behavior_type": "REQUIRED",
                "dataset_scope": ["ADSL"],
                "structure_scope": "SUBJECT_LEVEL_ANALYSIS_DATASET",
                "citation_provenance": citation,
                "citation_verification": verification,
                "semantic_review_status": "PENDING_APPROVAL",
                "promotion_authorized": False,
                "runtime_pack_admitted": False,
            }
    status = "FOUND_PENDING_APPROVAL" if direct_key_candidate else "KEY_EVIDENCE_UNRESOLVED"
    return {
        "schema_version": M3C_R2A_SCHEMA_VERSION,
        "status": status,
        "key_evidence_status": status,
        "search_scope": "Authorized local M3C-R1 ADSL candidate evidence and registry/source citation metadata.",
        "search_terms": search_terms,
        "inference_prohibitions": (
            "Do not infer formal key from one-record-per-subject.",
            "Do not infer formal key from populated USUBJID.",
            "Do not infer formal key from USUBJID uniqueness.",
        ),
        "candidate": direct_key_candidate,
        "reviewed_matches": reviewed_matches,
        "blocking_effect": (
            "Blocks key rule promotion, executable specification, and V1 READY only; does not block Knowledge "
            "Foundation completion or continued review of other citation-complete ADSL rules."
        ),
        "promotion_authorized": False,
        "runtime_pack_admitted": False,
    }


def _adsl_key_rejection_reason(text: str) -> str | None:
    lowered = text.lower()
    if "formal key" in lowered or "key variables" in lowered or "keys:" in lowered:
        return None
    if "one record per subject" in lowered:
        return "one-record-per-subject is grain evidence, not direct formal dataset key evidence"
    if "usubjid is present" in lowered or "values of usubjid are not present" in lowered:
        return "populated USUBJID is required-variable evidence, not direct formal dataset key evidence"
    if "unique value of usubjid" in lowered or "more than one record" in lowered:
        return "USUBJID uniqueness is duplicate-record evidence, not a complete formal dataset key"
    return "no direct formal ADSL key statement"


def _m3c_r2a_reconciliation(
    baseline_manifest: dict[str, Any],
    baseline_reconciliation: dict[str, Any],
    revisions: list[dict[str, Any]],
    key_search: dict[str, Any],
    active_total: int,
    resolved_total: int,
    unresolved_total: int,
) -> dict[str, Any]:
    return {
        "schema_version": M3C_R2A_SCHEMA_VERSION,
        "baseline_version": "m3c-r1",
        "baseline_artifact_hash": baseline_manifest["artifact_hash"],
        "baseline_counts": {
            "candidate_pack_count": baseline_manifest["candidate_pack_count"],
            "citation_resolved_candidate_count": baseline_manifest["citation_resolved_candidate_count"],
            "citation_unresolved_candidate_count": baseline_manifest["citation_unresolved_candidate_count"],
            "source_total": baseline_reconciliation.get("source_total"),
        },
        "adsl_parent_candidates_considered": 5,
        "adsl_revised_candidates_created": len(revisions),
        "one_for_one_replacements": len(revisions),
        "net_active_candidate_change_from_revisions": 0,
        "parent_records_deleted": 0,
        "parent_active_in_m3c_r2a": False,
        "revision_active_in_m3c_r2a": True,
        "key_evidence_status": key_search["key_evidence_status"],
        "active_candidate_total": active_total,
        "citation_resolved_candidate_count": resolved_total,
        "citation_unresolved_candidate_count": unresolved_total,
        "count_rules": {
            "parents_remain_only_in_historical_m3c_r1_lineage": True,
            "parent_and_revision_not_both_counted_in_m3c_r2a_active_pack": True,
            "key_evidence_unresolved_does_not_increase_active_candidate_total": key_search["candidate"] is None,
        },
        "gold_rules_written": 0,
        "runtime_rules_admitted": 0,
        "promotion_authorized": False,
        "thirty_day_teae_window_policy": {
            "classification": "STUDY_SPECIFIC",
            "automatic_standard_reclassification_allowed": False,
            "adae_behavior_implemented_in_m3c_r2a": False,
        },
    }


def _m3c_r2a_source_section_spot_checks(revisions: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": M3C_R2A_SCHEMA_VERSION,
        "selection_method": "M3C-R2A scoped five ADSL revisions; source-section classifications inherited from citation-verified parent evidence.",
        "OPERATIVE_RULE": [
            {
                "candidate_id": candidate["candidate_id"],
                "parent_candidate_id": candidate["parent_candidate_id"],
                "source_type": "OPERATIVE_RULE",
                "exact_source_excerpt": candidate["exact_source_excerpt"],
                "locator": candidate["citation_provenance"].get("locator"),
                "citation_verification_status": candidate["citation_verification"]["verification_status"],
                "classification_changed_by_audit": False,
            }
            for candidate in revisions
        ],
        "counts": {"OPERATIVE_RULE": len(revisions), "FRONT_MATTER": 0},
        "classifier_correction": {
            "publication_month_may_treated_as_normative_cue": False,
            "date_context_detection_required": True,
            "sentence_initial_normative_may_remains_normative_cue": True,
        },
    }


def _m3c_r2a_supersession_manifest(revisions: list[dict[str, Any]]) -> dict[str, Any]:
    old_artifacts = (
        "reviewer_a_results.json",
        "reviewer_b_results.json",
        "dual_review_comparison.json",
        "dual_review_comparison_corrected_v1.json",
        "dual_review_comparison_corrected_v2.json",
        "proposed_revised_candidates_v1.json",
        "proposed_revised_candidates_v2.json",
        "adjudication_batches/",
        "adjudication_batches_v2/",
        "phase_3e_immutability_check.json",
        "phase_3e_v2_immutability_check.json",
        "reconstructed_v1_core_review_queue.json",
        "reconstructed_review_input_freeze.json",
        "review_input_freeze.json",
    )
    return {
        "schema_version": M3C_R2A_SCHEMA_VERSION,
        "status": "SUPERSESSION_METADATA_ONLY",
        "supersession_scope": "M3C-R2A active ADSL review scope only",
        "parent_revision_lineage": [candidate["lineage"] for candidate in revisions],
        "superseded_review_artifacts": [
            {
                "artifact": artifact,
                "lineage_status": "SUPERSEDED_FOR_M3C_R2A_SCOPE",
                "preserved_unchanged": True,
            }
            for artifact in old_artifacts
        ],
        "promotion_authorized": False,
        "runtime_pack_admitted": False,
    }


def _extract_candidates(
    manifest: StandardManifest,
    source_path: Path,
    registry_root: Path,
) -> list[dict[str, Any]]:
    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_workbook_candidates(source_path)
    if suffix == ".xls":
        return _extract_legacy_workbook_candidates(source_path)
    return [
        {
            "text": block["text"],
            "locator": {"section": block.get("section"), "page": block.get("page")},
            "kind": _candidate_kind(str(block["text"])),
        }
        for block in _extract_text_blocks(source_path)
    ]


def _extract_workbook_candidates(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CandidateExtractionError("Workbook extraction requires openpyxl.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    candidates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        if "conformance" in path.name.lower() and sheet.title == "Rules Catalogue":
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = tuple(_cell_text(value) for value in rows[1])
            for row_index, row in enumerate(rows[2:], start=3):
                cells = tuple(_cell_text(value) for value in row)
                if not any(cells):
                    continue
                payload = {
                    _canonical_conformance_header(header) or f"column_{index + 1}": cells[index]
                    for index, header in enumerate(headers[: len(cells)])
                    if cells[index]
                }
                text = _conformance_rule_text(payload)
                candidates.append(
                    {
                        "text": text,
                        "locator": {
                            "sheet": sheet.title,
                            "row": row_index,
                            "table": "Rules Catalogue",
                            "official_rule_identifier": payload.get("rule_id"),
                        },
                        "kind": "conformance_rule",
                        "structured_fields": payload,
                        "source_section_classification": "OPERATIVE_RULE",
                    }
                )
            continue
        if "conformance" in path.name.lower():
            section_class = "REVISION_HISTORY" if "revision" in sheet.title.lower() or "retired" in sheet.title.lower() else "FRONT_MATTER"
            candidates.extend(_workbook_non_rule_rows(sheet, section_class))
            continue
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(_cell_text(value) for value in next(rows, ()))
        for row_index, row in enumerate(rows, start=2):
            cells = tuple(_cell_text(value) for value in row)
            if not any(cells):
                continue
            payload = {
                header or f"column_{index + 1}": cells[index]
                for index, header in enumerate(headers[: len(cells)])
                if cells[index]
            }
            text = " | ".join(f"{key}: {value}" for key, value in payload.items()) or " | ".join(cells)
            candidates.append(
                {
                    "text": text,
                    "locator": {"sheet": sheet.title, "row": row_index},
                    "kind": "conformance_rule" if "conformance" in path.name.lower() else "table_row",
                    "structured_fields": payload,
                    "source_section_classification": "OPERATIVE_RULE",
                }
            )
    return candidates


def _workbook_non_rule_rows(sheet: Any, source_section_classification: str) -> list[dict[str, Any]]:
    rows = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = tuple(_cell_text(value) for value in row)
        if not any(cells):
            continue
        text = " | ".join(value for value in cells if value)
        rows.append(
            {
                "text": text,
                "locator": {"sheet": sheet.title, "row": row_index},
                "kind": "non_rule_workbook_row",
                "source_section_classification": source_section_classification,
            }
        )
    return rows


def _extract_legacy_workbook_candidates(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise CandidateExtractionError("Legacy XLS extraction requires xlrd.") from exc

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    identity_markers = _legacy_workbook_identity_markers(workbook)
    if not any("ADaM" in marker or "Analysis Data Model" in marker for marker in identity_markers):
        raise CandidateExtractionError("Workbook identity could not be verified as ADaM Controlled Terminology.")
    candidates: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        headers = tuple(_cell_text(sheet.cell_value(0, col)) for col in range(sheet.ncols)) if sheet.nrows else ()
        for row_index in range(1, sheet.nrows):
            cells = tuple(_cell_text(sheet.cell_value(row_index, col)) for col in range(sheet.ncols))
            if not any(cells):
                continue
            payload = {
                headers[index] or f"column_{index + 1}": cells[index]
                for index in range(min(len(headers), len(cells)))
                if cells[index]
            }
            text = " | ".join(f"{key}: {value}" for key, value in payload.items()) or " | ".join(cells)
            candidates.append(
                {
                    "text": text,
                    "locator": {
                        "sheet": sheet.name,
                        "row": row_index + 1,
                        "workbook_identity_markers": identity_markers[:12],
                    },
                    "kind": "workbook_metadata"
                    if sheet.name.strip().lower() == "readme"
                    else "controlled_terminology",
                    "structured_fields": payload,
                    "source_section_classification": "FRONT_MATTER"
                    if sheet.name.strip().lower() == "readme"
                    else "CONTROLLED_TERMINOLOGY",
                }
            )
    workbook.release_resources()
    return candidates


def _legacy_workbook_identity_markers(workbook) -> tuple[str, ...]:
    markers: list[str] = []
    for sheet in workbook.sheets():
        for row_index in range(min(sheet.nrows, 20)):
            for col_index in range(min(sheet.ncols, 12)):
                value = _cell_text(sheet.cell_value(row_index, col_index))
                if value and any(token in value.lower() for token in ("adam", "analysis data model", "terminology", "publication", "version", "date")):
                    markers.append(value)
    return tuple(dict.fromkeys(markers))


def _normalize_candidates(manifest: StandardManifest, candidates: list[dict[str, Any]]) -> dict[str, Any]:
    retained_by_key: dict[str, dict[str, Any]] = {}
    removed_fragments: list[dict[str, Any]] = []
    merged_duplicates: list[dict[str, Any]] = []
    for candidate in candidates:
        text = str(candidate.get("text") or "").strip()
        normalized_text = _normalized_text(text)
        if candidate.get("kind") == "workbook_metadata":
            removed_fragments.append(
                {
                    "standard_id": manifest.id,
                    "reason": "workbook_metadata_not_controlled_terminology_rule",
                    "locator": candidate.get("locator"),
                    "excerpt": text[:240],
                }
            )
            continue
        source_section = _source_section_classification(text, manifest, candidate)
        if source_section in NON_OPERATIVE_SOURCE_SECTIONS:
            removed_fragments.append(
                {
                    "standard_id": manifest.id,
                    "reason": f"non_operative_source_section:{source_section}",
                    "locator": candidate.get("locator"),
                    "excerpt": text[:240],
                }
            )
            continue
        if _is_fragment_or_header(normalized_text):
            removed_fragments.append(
                {
                    "standard_id": manifest.id,
                    "reason": "fragment_or_repeated_header",
                    "locator": candidate.get("locator"),
                    "excerpt": text[:240],
                }
            )
            continue
        locator_key = _locator_key(candidate.get("locator") or {})
        if manifest.id == "adam-conformance-rules":
            stable_key = f"{manifest.id}|{locator_key}"
        else:
            stable_key = f"{manifest.id}|{_hash_text(normalized_text)}"
        if stable_key in retained_by_key:
            retained_by_key[stable_key].setdefault("duplicate_locators", []).append(candidate.get("locator"))
            merged_duplicates.append(
                {
                    "standard_id": manifest.id,
                    "reason": "duplicate_normalized_requirement",
                    "retained_locator": retained_by_key[stable_key].get("locator"),
                    "duplicate_locator": candidate.get("locator"),
                }
            )
            continue
        item = dict(candidate)
        item["source_section_classification"] = source_section
        item["stable_key"] = stable_key
        item["source_excerpt"] = text
        item["normalized_text_hash"] = _hash_text(normalized_text)
        item["duplicate_locators"] = []
        retained_by_key[stable_key] = item
    return {
        "retained": [retained_by_key[key] for key in sorted(retained_by_key)],
        "removed_fragments": removed_fragments,
        "merged_duplicates": merged_duplicates,
    }


def _extract_text_blocks(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return _extract_plain_text_blocks(path)
    if suffix == ".pdf":
        return _extract_pdf_blocks(path)
    raise CandidateExtractionError(f"Unsupported standards document type: {suffix}")


def _extract_plain_text_blocks(path: Path) -> list[dict[str, Any]]:
    content = path.read_text(encoding="utf-8")
    blocks: list[dict[str, Any]] = []
    section: str | None = None
    page: int | None = None
    paragraph: list[str] = []

    def flush() -> None:
        nonlocal paragraph
        text = " ".join(line.strip() for line in paragraph if line.strip())
        if text:
            blocks.append({"text": text, "section": section, "page": page})
        paragraph = []

    for raw_line in content.splitlines():
        line = raw_line.strip()
        if not line:
            flush()
            continue
        page_match = re.fullmatch(r"\[page\s+([0-9]+)\]", line, flags=re.IGNORECASE)
        if page_match:
            flush()
            page = int(page_match.group(1))
            continue
        heading_match = re.match(r"^(?:#{1,6}\s+|section\s+)(.+)$", line, flags=re.IGNORECASE)
        if heading_match:
            flush()
            section = heading_match.group(1).strip()
            continue
        paragraph.append(line)
    flush()
    return blocks


def _extract_pdf_blocks(path: Path) -> list[dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise CandidateExtractionError("PDF extraction requires pypdf.") from exc

    try:
        reader = PdfReader(str(path))
        blocks: list[dict[str, Any]] = []
        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for paragraph in re.split(r"\n\s*\n", text):
                normalized = re.sub(r"\s+", " ", paragraph).strip()
                if normalized:
                    blocks.append({"text": normalized, "section": None, "page": page_index})
        return blocks
    except Exception as exc:
        raise CandidateExtractionError(f"Failed to extract PDF text from {path}.") from exc


def _candidate_rule(
    manifest: StandardManifest,
    source_path: Path,
    registry_root: Path,
    candidate: dict[str, Any],
    index: int,
    extracted_at: str,
) -> CompiledRule:
    text = str(candidate["text"])
    role = manifest.role
    source_section = _source_section_classification(text, manifest, candidate)
    exact_excerpt = str(candidate.get("source_excerpt") or text).strip()
    structured_fields = dict(candidate.get("structured_fields") or {})
    normalized_requirement = _normalized_atomic_requirement(text, manifest, candidate)
    normative_strength = _normative_strength(text, role, source_section, candidate["kind"])
    classification = _classification(normative_strength, role, source_section, candidate["kind"])
    locator = dict(candidate["locator"])
    locator["source_excerpt"] = _short_excerpt(exact_excerpt)
    locator["duplicate_locators"] = tuple(candidate.get("duplicate_locators") or ())
    locator["source_binding_status"] = _source_binding_status(locator, manifest)
    locator["extraction_timestamp"] = extracted_at
    locator["parser_compiler_version"] = KNOWLEDGE_COMPILER_VERSION
    classification_status = "RESOLVED" if classification is not None and locator["source_binding_status"] == "SOURCE_BOUND" else "UNRESOLVED"
    dataset_scope = _dataset_scope(text, structured_fields)
    variable_scope = _variable_scope(text, structured_fields, dataset_scope)
    structure_scope = _structure_scope(text, structured_fields, dataset_scope)
    validation_requirement = _validation_requirement_text(text, manifest, structured_fields)
    required_behavior = _required_behavior_text(text, role, source_section, candidate["kind"], structured_fields)
    return CompiledRule(
        rule_id=f"{manifest.id}:{candidate.get('stable_key', str(index)).split('|')[-1]}",
        standard_id=manifest.id,
        standard_name=manifest.title,
        standard_version=manifest.version,
        classification=classification,
        scope=_scope(text, candidate["kind"], dataset_scope=dataset_scope, variable_scope=variable_scope, structure_scope=structure_scope),
        applicability_conditions=_conditions(text),
        required_behavior=required_behavior,
        permitted_variations=tuple(_short_excerpt(item) for item in _sentences_matching(text, ("may", "can", "optional", "should"))),
        prohibited_behavior=tuple(_short_excerpt(item) for item in _sentences_matching(text, ("must not", "shall not", "prohibited", "not allowed"))),
        required_inputs=_source_references(text),
        expected_outputs=tuple(sorted(set(dataset_scope) | set(f"{dataset}.{variable}" for dataset in dataset_scope for variable in variable_scope))),
        grain=_grain(text),
        keys=_keys(text),
        validation_requirements=_validation_requirements(text, manifest),
        evidence_locator=locator,
        local_relative_source_path=_relative_source(registry_root, source_path),
        official_url=manifest.official_url,
        source_hash=manifest.sha256,
        extraction_status="CANDIDATE_RULE",
        review_status="CANDIDATE",
        classification_status=classification_status,
        normative_strength=normative_strength,
        source_section_classification=source_section,
        exact_source_excerpt=exact_excerpt,
        normalized_atomic_requirement=normalized_requirement,
        rule_type=_rule_type(candidate["kind"], source_section, manifest),
        dataset_scope=dataset_scope,
        structure_scope=structure_scope,
        variable_scope=variable_scope,
        validation_requirement=validation_requirement,
        semantic_reconstruction_status="CANDIDATE_RECONSTRUCTED",
        excluded_reason=None,
        conformance_metadata=_conformance_metadata(structured_fields) if manifest.id == "adam-conformance-rules" else None,
    )


def _status_rule(
    manifest: StandardManifest,
    integrity: dict[str, Any],
    extracted_at: str,
) -> CompiledRule:
    return CompiledRule(
        rule_id=f"{manifest.id}:status",
        standard_id=manifest.id,
        standard_name=manifest.title,
        standard_version=manifest.version,
        classification=None,
        scope={"source_role": manifest.role},
        applicability_conditions=(),
        required_behavior=None,
        permitted_variations=(),
        prohibited_behavior=(),
        required_inputs=(),
        expected_outputs=(),
        grain=None,
        keys=(),
        validation_requirements=(),
        evidence_locator={
            "status": integrity["status"],
            "reason": integrity.get("reason"),
            "extraction_timestamp": extracted_at,
            "parser_compiler_version": KNOWLEDGE_COMPILER_VERSION,
        },
        local_relative_source_path=manifest.local_path,
        official_url=manifest.official_url,
        source_hash=manifest.sha256,
        extraction_status=integrity["status"],
        review_status="CANDIDATE",
        classification_status="UNRESOLVED",
        normative_strength=None,
        source_section_classification="UNRESOLVED",
        exact_source_excerpt=None,
        normalized_atomic_requirement=None,
        rule_type=None,
        dataset_scope=(),
        structure_scope=(),
        variable_scope=(),
        validation_requirement=None,
        semantic_reconstruction_status="SOURCE_STATUS_ONLY",
        excluded_reason=integrity["status"],
        conformance_metadata=None,
    )


def _classification(normative_strength: str, role: str, source_section: str = "UNRESOLVED", kind: str = "") -> str | None:
    if role == "validation_reference":
        return "NON_NORMATIVE"
    if role == "upstream_reference":
        return "NON_NORMATIVE"
    if source_section not in NORMATIVE_SOURCE_SECTIONS and kind != "conformance_rule":
        return None
    if normative_strength == "REQUIRED":
        return "STANDARD_REQUIRED"
    if normative_strength in {"GUIDED", "PERMITTED"}:
        return "STANDARD_GUIDED"
    return None


def _normative_strength(text: str, role: str, source_section: str = "UNRESOLVED", kind: str = "") -> str:
    lowered = text.lower()
    if role != "primary_standard":
        return "NON_NORMATIVE"
    if source_section not in NORMATIVE_SOURCE_SECTIONS and kind != "conformance_rule":
        return "NON_NORMATIVE"
    if kind == "conformance_rule":
        return "REQUIRED"
    if any(token in lowered for token in ("must ", "shall ", "required", "not permitted", "must not", "shall not")):
        return "REQUIRED"
    if any(_contains_token(text, token) for token in ("should ", "recommended", " may ", "can ", "optional")):
        return "GUIDED"
    return "UNRESOLVED"


def _scope(
    text: str,
    kind: str,
    *,
    dataset_scope: tuple[str, ...] | None = None,
    variable_scope: tuple[str, ...] | None = None,
    structure_scope: tuple[str, ...] | None = None,
) -> dict[str, Any]:
    datasets = dataset_scope if dataset_scope is not None else _datasets(text)
    variables = variable_scope if variable_scope is not None else _variable_scope(text, {}, datasets)
    return {
        "candidate_kind": kind,
        "datasets": datasets,
        "variables": variables,
        "dataset_scope": datasets,
        "structure_scope": structure_scope if structure_scope is not None else _structure_scope(text, {}, datasets),
        "variable_scope": variables,
    }


def _datasets(text: str) -> tuple[str, ...]:
    return tuple(dataset for dataset in V1_DATASETS if re.search(rf"\b{dataset}\b", text, flags=re.IGNORECASE))


def _adam_outputs(text: str) -> tuple[str, ...]:
    outputs = set()
    for match in re.finditer(r"\b(ADSL|ADAE|ADLB|ADTTE)\.([A-Z][A-Z0-9_]{1,31})\b", text):
        outputs.add(f"{match.group(1)}.{match.group(2)}")
    for dataset in V1_DATASETS:
        if re.search(rf"\b{dataset}\b", text):
            outputs.add(dataset)
    return tuple(sorted(outputs))


def _source_references(text: str) -> tuple[str, ...]:
    refs = set()
    for match in re.finditer(r"\b(DM|AE|LB|DS|EX|SV)\.([A-Z][A-Z0-9_]{1,31})\b", text):
        refs.add(f"{match.group(1)}.{match.group(2)}")
    return tuple(sorted(refs))


def _conditions(text: str) -> tuple[str, ...]:
    return _sentences_matching(text, (" if ", " when ", " where ", " unless ", " except "))


def _validation_requirements(text: str, manifest: StandardManifest) -> tuple[str, ...]:
    requirements = []
    if manifest.id == "adam-conformance-rules":
        requirements.append("Conformance rule row requires independent evaluation before runtime use.")
    requirements.extend(_sentences_matching(text, ("validate", "validation", "conformance", "must", "shall")))
    return tuple(dict.fromkeys(requirements))


def _grain(text: str) -> str | None:
    match = re.search(r"one record per[^.;:\n]+", text, flags=re.IGNORECASE)
    return match.group(0).strip() if match else None


def _keys(text: str) -> tuple[str, ...]:
    candidates = []
    for match in re.finditer(r"\b(?:key|keys|identifier|identifiers)\b[^.;:\n]*", text, flags=re.IGNORECASE):
        candidates.extend(re.findall(r"\b[A-Z][A-Z0-9_]{1,31}\b", match.group(0)))
    return tuple(dict.fromkeys(candidates))


def _operator_hint(text: str) -> str | None:
    lowered = text.lower()
    if any(token in lowered for token in ("sum", "mean", "average", "count")):
        return "aggregate"
    if any(token in lowered for token in ("join", "merge")):
        return "join"
    if "baseline" in lowered:
        return "select_baseline"
    if "date" in lowered and any(token in lowered for token in ("difference", "duration", "elapsed")):
        return "date_difference"
    if any(token in lowered for token in ("first", "earliest")):
        return "select_first"
    if any(token in lowered for token in ("last", "latest")):
        return "select_last"
    if any(token in lowered for token in ("numeric", "integer", "decimal")):
        return "numeric_conversion"
    if "date" in lowered:
        return "parse_date"
    return None


def _candidate_kind(text: str) -> str:
    lowered = text.lower()
    if "controlled terminology" in lowered or "codelist" in lowered or "submission value" in lowered:
        return "controlled_terminology"
    if "variable" in lowered or "parameter" in lowered:
        return "variable_requirement"
    if "one record per" in lowered or "structure" in lowered:
        return "dataset_structure"
    if "conformance" in lowered or "rule" in lowered:
        return "conformance_rule"
    return "prose_requirement"


def _canonical_conformance_header(header: str) -> str:
    normalized = re.sub(r"\s+", " ", header).strip().lower()
    mapping = {
        "rule id": "rule_id",
        "rule id version (represents any change to the rule)": "rule_version",
        "related rule(s)": "related_rules",
        "rule set (generally ig version, occds v1.0, adnca v1.0)": "rule_set",
        "class": "class",
        "subclass": "subclass",
        "send/sdtm domain": "source_domain",
        "variable or item": "variable_or_item",
        "define-xml element": "define_xml_element",
        "scope section": "scope_section",
        "natural language rule (success criteria)": "success_message",
        "rule (success criteria)": "success_rule",
        "condition (success)": "success_condition",
        "natural language rule (failure criteria)": "failure_message",
        "rule (failure criteria)": "failure_rule",
        "condition (failure)": "failure_condition",
        "rule section": "rule_section",
        "implementation guide (cited document)": "cited_standard",
        "cited section": "cited_section",
        "cited item (text; figure; table; footnote)": "cited_item",
        "cited guidance": "cited_guidance",
        "guidance section": "guidance_section",
        "release notes": "release_notes",
    }
    return mapping.get(normalized, normalized.replace(" ", "_").replace("/", "_").replace("-", "_"))


def _conformance_rule_text(payload: dict[str, str]) -> str:
    parts = [
        ("Rule ID", payload.get("rule_id")),
        ("Rule Version", payload.get("rule_version")),
        ("Rule Set", payload.get("rule_set")),
        ("Scope", payload.get("class")),
        ("Subclass", payload.get("subclass")),
        ("Variable or Item", payload.get("variable_or_item")),
        ("Condition", payload.get("failure_condition") or payload.get("success_condition")),
        ("Error Message", payload.get("failure_message") or payload.get("success_message")),
        ("Rule", payload.get("failure_rule") or payload.get("success_rule")),
        ("Cited Standard", payload.get("cited_standard")),
        ("Cited Section", payload.get("cited_section")),
        ("Cited Item", payload.get("cited_item")),
        ("Cited Guidance", payload.get("cited_guidance")),
    ]
    return " | ".join(f"{name}: {value}" for name, value in parts if value)


def _source_section_classification(
    text: str,
    manifest: StandardManifest,
    candidate: dict[str, Any] | None = None,
) -> str:
    candidate = candidate or {}
    explicit = candidate.get("source_section_classification")
    if explicit in SOURCE_SECTION_CLASSIFICATIONS:
        return str(explicit)
    kind = str(candidate.get("kind") or "")
    lowered = text.lower()
    locator = candidate.get("locator") or {}
    sheet = str(locator.get("sheet") or "").lower()
    if kind == "controlled_terminology":
        return "CONTROLLED_TERMINOLOGY"
    if kind == "conformance_rule" and manifest.id == "adam-conformance-rules" and sheet == "rules catalogue":
        return "OPERATIVE_RULE"
    if "revision history" in lowered or "retired rules" in lowered or "deprecated" in lowered:
        return "REVISION_HISTORY"
    if "glossary" in lowered or "appendix a: glossary" in lowered:
        return "GLOSSARY"
    if "example" in lowered or "sample " in lowered or "for illustration only" in lowered:
        return "EXAMPLE"
    if "copyright" in lowered or "table of contents" in lowered or "introduction / purpose" in lowered:
        return "FRONT_MATTER"
    if "ectd folder" in lowered or "submission" in lowered or "future adam data structures" in lowered:
        return "SUBMISSION_CONTEXT"
    if kind in {"dataset_structure", "variable_requirement", "conformance_rule", "prose_requirement"}:
        if _contains_normative_cue(text):
            return "OPERATIVE_RULE"
        if "definition" in lowered or "defined as" in lowered:
            return "DEFINITION"
    return "UNRESOLVED"


def _dataset_scope(text: str, structured_fields: dict[str, Any]) -> tuple[str, ...]:
    candidates = set(_datasets(text))
    for field in ("class", "subclass", "scope_section", "variable_or_item"):
        value = str(structured_fields.get(field) or "")
        candidates.update(_datasets(value))
        if value.upper() == "SUBJECT LEVEL ANALYSIS DATASET":
            candidates.add("ADSL")
    return tuple(sorted(candidates))


def _variable_scope(text: str, structured_fields: dict[str, Any], dataset_scope: tuple[str, ...]) -> tuple[str, ...]:
    values = set()
    for match in re.finditer(r"\b(?:ADSL|ADAE|ADLB|ADTTE)\.([A-Z][A-Z0-9_]{1,31})\b", text):
        values.add(match.group(1))
    variable_or_item = str(structured_fields.get("variable_or_item") or "")
    for token in re.findall(r"\b[A-Z][A-Z0-9_]{1,31}\b", variable_or_item):
        if token not in V1_DATASETS and token not in {"ALL"}:
            values.add(token)
    if not values:
        for token in re.findall(r"\b[A-Z][A-Z0-9_]{2,31}\b", text):
            if token not in V1_DATASETS and token not in {"CDISC", "ADAM", "SDTM", "SEND", "OCCDS", "BDS"}:
                values.add(token)
    return tuple(sorted(values))


def _structure_scope(text: str, structured_fields: dict[str, Any], dataset_scope: tuple[str, ...]) -> tuple[str, ...]:
    lowered = text.lower()
    structures = set()
    class_value = str(structured_fields.get("class") or "")
    if class_value:
        structures.add(class_value)
    if "one record per subject" in lowered or "subject-level analysis dataset" in lowered:
        structures.add("SUBJECT_LEVEL_ANALYSIS_DATASET")
    if "occurrence data structure" in lowered or "occds" in lowered:
        structures.add("OCCURRENCE_DATA_STRUCTURE")
    if "basic data structure" in lowered or " bds" in lowered:
        structures.add("BASIC_DATA_STRUCTURE")
    return tuple(sorted(structures))


def _normalized_atomic_requirement(text: str, manifest: StandardManifest, candidate: dict[str, Any]) -> str | None:
    structured = dict(candidate.get("structured_fields") or {})
    if manifest.id == "adam-conformance-rules":
        return _short_excerpt(
            structured.get("failure_message")
            or structured.get("success_message")
            or structured.get("failure_rule")
            or structured.get("success_rule")
            or text,
            limit=360,
        )
    sentences = _sentences_matching(text, NORMATIVE_CUE_TOKENS)
    return _short_excerpt(sentences[0] if sentences else text, limit=360)


def _required_behavior_text(
    text: str,
    role: str,
    source_section: str,
    kind: str,
    structured_fields: dict[str, Any],
) -> str | None:
    if role != "primary_standard":
        return None
    if source_section not in NORMATIVE_SOURCE_SECTIONS and kind != "conformance_rule":
        return None
    if kind == "conformance_rule":
        return _short_excerpt(
            structured_fields.get("failure_rule")
            or structured_fields.get("success_rule")
            or structured_fields.get("failure_message")
            or structured_fields.get("success_message")
            or text,
            limit=360,
        )
    sentences = _sentences_matching(text, NORMATIVE_CUE_TOKENS)
    return _short_excerpt(sentences[0] if sentences else text, limit=360)


def _validation_requirement_text(text: str, manifest: StandardManifest, structured_fields: dict[str, Any]) -> str | None:
    if manifest.id == "adam-conformance-rules":
        return _short_excerpt(structured_fields.get("failure_message") or structured_fields.get("success_message") or text, limit=360)
    matches = _validation_requirements(text, manifest)
    return _short_excerpt(matches[0], limit=360) if matches else None


def _rule_type(kind: str, source_section: str, manifest: StandardManifest) -> str:
    if manifest.id == "adam-conformance-rules":
        return "CONFORMANCE_RULE"
    if source_section == "CONTROLLED_TERMINOLOGY":
        return "CONTROLLED_TERMINOLOGY"
    if source_section == "DEFINITION":
        return "DEFINITION"
    return "OPERATIVE_RULE" if source_section == "OPERATIVE_RULE" else "NON_OPERATIVE_REFERENCE"


def _conformance_metadata(structured_fields: dict[str, Any]) -> dict[str, Any]:
    return {
        "official_rule_identifier": structured_fields.get("rule_id"),
        "rule_version": structured_fields.get("rule_version"),
        "related_rules": structured_fields.get("related_rules"),
        "rule_set": structured_fields.get("rule_set"),
        "scope": {
            "class": structured_fields.get("class"),
            "subclass": structured_fields.get("subclass"),
            "source_domain": structured_fields.get("source_domain"),
            "variable_or_item": structured_fields.get("variable_or_item"),
            "define_xml_element": structured_fields.get("define_xml_element"),
            "scope_section": structured_fields.get("scope_section"),
        },
        "condition": structured_fields.get("failure_condition") or structured_fields.get("success_condition"),
        "error_message": structured_fields.get("failure_message") or structured_fields.get("success_message"),
        "rule_expression": structured_fields.get("failure_rule") or structured_fields.get("success_rule"),
        "referenced_standard": structured_fields.get("cited_standard"),
        "referenced_section": structured_fields.get("cited_section"),
        "referenced_item": structured_fields.get("cited_item"),
        "cited_guidance": structured_fields.get("cited_guidance"),
        "guidance_section": structured_fields.get("guidance_section"),
        "release_notes": structured_fields.get("release_notes"),
    }


def _sentences_matching(text: str, tokens: tuple[str, ...]) -> tuple[str, ...]:
    sentences = re.split(r"(?<=[.!?])\s+", text)
    matches = []
    for sentence in sentences:
        if any(_contains_token(sentence, token) for token in tokens):
            matches.append(sentence.strip())
    return tuple(dict.fromkeys(item for item in matches if item))


def _contains_normative_cue(text: str) -> bool:
    return any(_contains_token(text, token) for token in NORMATIVE_CUE_TOKENS)


def _contains_token(text: str, token: str) -> bool:
    normalized = token.strip().lower()
    if normalized == "may":
        return _first_modal_may_match(text) is not None
    return token.lower() in text.lower()


def _first_normative_cue(text: str) -> str | None:
    for token in NORMATIVE_CUE_TOKENS:
        if _contains_token(text, token):
            return token.strip()
    return None


def _normative_cue_index(text: str, token: str) -> int:
    normalized = token.strip().lower()
    if normalized == "may":
        match = _first_modal_may_match(text)
        return match.start() if match else -1
    return text.lower().find(token.lower())


def _first_modal_may_match(text: str) -> re.Match[str] | None:
    for match in re.finditer(r"(?<![A-Za-z])may(?![A-Za-z])", text, flags=re.IGNORECASE):
        if not MONTH_DATE_PATTERN.match(text, match.start()):
            return match
    return None


def _dataset_structures(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    rows = []
    for dataset in V1_DATASETS:
        supporting = _supporting_rules(rules, dataset, ("dataset_structure",))
        rows.append(_coverage_row(dataset, "dataset_structure", supporting, study_decision_required=False))
        rows.append(_coverage_row(dataset, "grain", tuple(rule for rule in supporting if rule.grain), study_decision_required=False))
        rows.append(_coverage_row(dataset, "keys", tuple(rule for rule in supporting if rule.keys), study_decision_required=False))
    return rows


def _variable_requirements(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    rows = []
    for dataset in V1_DATASETS:
        supporting = _supporting_rules(rules, dataset, ("variable_requirement", "table_row"))
        rows.append(_coverage_row(dataset, "required_or_conditional_variables", supporting, study_decision_required=False))
        rows.append(_coverage_row(dataset, "source_traceability", tuple(rule for rule in supporting if rule.required_inputs), study_decision_required=False))
        rows.append(_coverage_row(dataset, "derivation_constraints", supporting, study_decision_required=True))
        rows.append(_coverage_row(dataset, "missing_value_constraints", supporting, study_decision_required=True))
    return rows


def _conformance_rules(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    return [
        {
            "rule_id": rule.rule_id,
            "standard_id": rule.standard_id,
            "classification": rule.classification,
            "validation_requirements": rule.validation_requirements,
            "citation_status": _citation_status(rule),
            "review_status": rule.review_status,
        }
        for rule in rules
        if rule.standard_id == "adam-conformance-rules" and rule.extraction_status == "CANDIDATE_RULE"
    ]


def _rule_precedence(rules: tuple[CompiledRule, ...]) -> dict[str, Any]:
    conflicts = []
    for dataset in V1_DATASETS:
        by_item: dict[str, set[str]] = {}
        for rule in _supporting_rules(rules, dataset, ("dataset_structure", "variable_requirement", "table_row", "prose_requirement")):
            for output in rule.expected_outputs or (dataset,):
                if rule.classification is not None:
                    by_item.setdefault(output, set()).add(rule.classification)
        conflicts.extend(
            {
                "scope": scope,
                "classifications": sorted(classifications),
                "status": "REVIEW_REQUIRED",
            }
            for scope, classifications in sorted(by_item.items())
            if "STANDARD_REQUIRED" in classifications and len(classifications) > 1
        )
    conflicts = list({
        (item["scope"], tuple(item["classifications"])): item
        for item in conflicts
    }.values())
    by_scope = {item["scope"]: item for item in conflicts}
    for scope in CONFLICT_LINEAGE_SCOPES:
        by_scope.setdefault(
            scope,
            {
                "scope": scope,
                "classifications": ["STANDARD_GUIDED", "STANDARD_REQUIRED"],
                "status": "REVIEW_REQUIRED",
                "lineage_source": "phase_3b_detected_conflict",
            },
        )
    conflicts = [
        {
            "conflict_id": f"CONFLICT-LINEAGE-{index:03d}",
            **by_scope[scope],
        }
        for index, scope in enumerate(CONFLICT_LINEAGE_SCOPES, start=1)
    ]
    return {
        "order": PRECEDENCE,
        "rules": [
            "STANDARD_REQUIRED cannot be overridden.",
            "STANDARD_GUIDED may allow documented study-specific variation only when source evidence permits it.",
            "Cross-standard conflicts are recorded for review rather than silently resolved.",
        ],
        "detected_conflicts": conflicts,
        "review_status": "CANDIDATE",
    }


def _coverage_matrix(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    rows = []
    for dataset in V1_DATASETS:
        rows.extend(
            [
                _coverage_row(dataset, "dataset_structure", _supporting_rules(rules, dataset, ("dataset_structure",)), study_decision_required=False),
                _coverage_row(dataset, "grain", tuple(rule for rule in _supporting_rules(rules, dataset, ("dataset_structure", "table_row")) if rule.grain), study_decision_required=False),
                _coverage_row(dataset, "keys", tuple(rule for rule in _supporting_rules(rules, dataset, ("dataset_structure", "table_row")) if rule.keys), study_decision_required=False),
                _coverage_row(dataset, "required_or_conditional_variables", _supporting_rules(rules, dataset, ("variable_requirement", "table_row")), study_decision_required=False),
                _coverage_row(dataset, "source_traceability", tuple(rule for rule in _supporting_rules(rules, dataset, ("variable_requirement", "table_row", "prose_requirement")) if rule.required_inputs), study_decision_required=False),
                _coverage_row(dataset, "derivation_constraints", _supporting_rules(rules, dataset, ("prose_requirement", "table_row")), study_decision_required=True),
                _coverage_row(dataset, "missing_value_constraints", _supporting_rules(rules, dataset, ("prose_requirement", "table_row")), study_decision_required=True),
                _coverage_row(dataset, "validation_conformance_requirements", tuple(rule for rule in _supporting_rules(rules, dataset, ("conformance_rule", "table_row")) if rule.validation_requirements), study_decision_required=False),
                _coverage_row(dataset, "citation_availability", tuple(rule for rule in _supporting_rules(rules, dataset, ("dataset_structure", "variable_requirement", "prose_requirement", "table_row")) if _citation_status(rule) == "RESOLVED"), study_decision_required=False),
            ]
        )
    return rows


def _coverage_row(
    dataset: str,
    item: str,
    supporting_rules: tuple[CompiledRule, ...],
    *,
    study_decision_required: bool,
) -> dict[str, Any]:
    unresolved = []
    if not supporting_rules:
        unresolved.append("No candidate rule with resolved citation supports this item.")
    if any(_citation_status(rule) != "RESOLVED" for rule in supporting_rules):
        unresolved.append("One or more supporting citations are unresolved.")
    if any(rule.classification_status != "RESOLVED" for rule in supporting_rules):
        unresolved.append("One or more supporting rule classifications are unresolved.")
    return {
        "dataset": dataset,
        "coverage_item": item,
        "supporting_rule_ids": tuple(rule.rule_id for rule in supporting_rules),
        "source_documents": tuple(sorted({rule.standard_id for rule in supporting_rules})),
        "coverage_status": "INCOMPLETE",
        "unresolved_conflicts": tuple(unresolved + ["Semantic review and independent reviewer decision are pending."]),
        "study_decision_required": study_decision_required,
        "review_status": "CANDIDATE",
    }


def _citation_index(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    citations = []
    for rule in rules:
        citations.append(
            {
                "rule_id": rule.rule_id,
                "standard_name": rule.standard_name,
                "standard_version": rule.standard_version,
                "local_relative_path": rule.local_relative_source_path,
                "official_url": rule.official_url,
                "locator": rule.evidence_locator,
                "source_hash": rule.source_hash,
                "citation_status": _citation_status(rule),
                "review_status": rule.review_status,
            }
        )
    return citations


def _review_queue(rules: tuple[CompiledRule, ...], *, limit: int = 180) -> list[dict[str, Any]]:
    candidates = [
        rule for rule in rules
        if rule.extraction_status == "CANDIDATE_RULE"
        and rule.review_status == "CANDIDATE"
        and rule.classification in {"STANDARD_REQUIRED", "STANDARD_GUIDED", "NON_NORMATIVE"}
        and _rule_datasets(rule)
    ]
    scored = [
        rule for rule in sorted(
            candidates,
            key=lambda rule: (
                -_review_priority(rule),
                rule.standard_id,
                rule.rule_id,
            ),
        )
    ]
    selected: list[CompiledRule] = []
    selected_ids: set[str] = set()
    dataset_counts: dict[str, int] = {}
    family_counts: dict[tuple[str, str], int] = {}
    target_families = (
        "structure_grain_keys",
        "required_or_conditional_variables",
        "traceability",
        "derivation_constraints",
        "missing_value_constraints",
        "validation_conformance_requirements",
        "controlled_terminology",
    )
    for dataset in V1_DATASETS:
        for family in target_families:
            matches = [
                rule for rule in scored
                if rule.rule_id not in selected_ids
                and dataset in _rule_datasets(rule)
                and _rule_family(rule) == family
            ][:4]
            for rule in matches:
                selected.append(rule)
                selected_ids.add(rule.rule_id)
                for rule_dataset in _rule_datasets(rule):
                    dataset_counts[rule_dataset] = dataset_counts.get(rule_dataset, 0) + 1
                    key = (rule_dataset, _rule_family(rule))
                    family_counts[key] = family_counts.get(key, 0) + 1
    for rule in scored:
        if len(selected) >= limit:
            break
        if rule.rule_id in selected_ids:
            continue
        datasets = _rule_datasets(rule)
        family = _rule_family(rule)
        if all(dataset_counts.get(dataset, 0) >= 70 for dataset in datasets):
            continue
        if all(family_counts.get((dataset, family), 0) >= 18 for dataset in datasets):
            continue
        selected.append(rule)
        selected_ids.add(rule.rule_id)
        for dataset in datasets:
            dataset_counts[dataset] = dataset_counts.get(dataset, 0) + 1
            key = (dataset, family)
            family_counts[key] = family_counts.get(key, 0) + 1
    selected = sorted(
        selected[:limit],
        key=lambda rule: (
            -_review_priority(rule),
            rule.standard_id,
            rule.rule_id,
        ),
    )
    return [_review_item(rule, rules) for rule in selected]


def _semantic_review_queue(rules: tuple[CompiledRule, ...], *, limit: int = 180) -> list[dict[str, Any]]:
    normative_rules = tuple(
        rule
        for rule in rules
        if rule.classification in {"STANDARD_REQUIRED", "STANDARD_GUIDED"}
        and rule.classification_status == "RESOLVED"
        and _citation_provenance_complete(rule)
        and rule.source_section_classification in NORMATIVE_SOURCE_SECTIONS
    )
    return _review_queue(normative_rules, limit=limit)


def _review_item(rule: CompiledRule, rules: tuple[CompiledRule, ...]) -> dict[str, Any]:
    related = [
        other.rule_id for other in rules
        if other.rule_id != rule.rule_id
        and other.extraction_status == "CANDIDATE_RULE"
        and set(_rule_datasets(other)) & set(_rule_datasets(rule))
        and _rule_family(other) == _rule_family(rule)
    ][:10]
    return {
        "candidate_rule_id": rule.rule_id,
        "rule_family": _rule_family(rule),
        "standard": rule.standard_name,
        "standard_version": rule.standard_version,
        "source_role": _source_role(rule),
        "exact_locator": _locator_without_long_fields(rule.evidence_locator),
        "short_source_excerpt": rule.evidence_locator.get("source_excerpt"),
        "proposed_classification": rule.classification,
        "proposed_structured_requirement": {
            "normative_strength": rule.normative_strength,
            "classification_status": rule.classification_status,
            "required_behavior": rule.required_behavior,
            "validation_requirements": rule.validation_requirements,
        },
        "applicability_conditions": rule.applicability_conditions,
        "dataset_variable_scope": rule.scope,
        "required_behavior": rule.required_behavior,
        "permitted_behavior": rule.permitted_variations,
        "prohibited_behavior": rule.prohibited_behavior,
        "related_candidate_rules": related,
        "possible_conflict_ids": (),
        "source_binding_status": rule.evidence_locator.get("source_binding_status"),
        "semantic_review_status": "PENDING",
        "reviewer_decision": None,
    }


def _conflict_review(
    rules: tuple[CompiledRule, ...],
    conflicts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = []
    grouped: dict[str, list[dict[str, Any]]] = {}
    for conflict in conflicts:
        dataset = str(conflict["scope"]).split(".", 1)[0]
        grouped.setdefault(dataset, []).append(conflict)
    for index, dataset in enumerate(("ADAE", "ADSL", "ADTTE", "ADLB"), start=1):
        group = grouped.get(dataset, ())
        scopes = tuple(conflict["scope"] for conflict in group)
        classifications = {
            classification
            for conflict in group
            for classification in conflict.get("classifications", ())
        }
        related = [
            rule for rule in rules
            if rule.extraction_status == "CANDIDATE_RULE"
            and rule.classification in classifications
            and (
                dataset in rule.scope.get("datasets", ())
                or any(scope in rule.expected_outputs or scope in rule.scope.get("variables", ()) for scope in scopes)
            )
        ][:8]
        rows.append(
            {
                "conflict_id": f"CONFLICT-{index:03d}",
                "scope": dataset,
                "underlying_conflict_ids": tuple(conflict["conflict_id"] for conflict in group),
                "underlying_scopes": scopes,
                "rule_ids": tuple(rule.rule_id for rule in related),
                "source_standards": tuple(
                    {
                        f"{rule.standard_id} {rule.standard_version or ''}".strip()
                        for rule in related
                    }
                ),
                "requirement_summaries": tuple(_short_excerpt(rule.required_behavior or rule.evidence_locator.get("source_excerpt", "")) for rule in related),
                "proposed_assessment": _conflict_assessment(dataset, related),
                "review_status": "PENDING",
            }
        )
    return rows


def _coverage_gap_review_map(
    coverage_rows: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = []
    for row in coverage_rows:
        if row["coverage_status"] != "INCOMPLETE":
            continue
        dataset = row["dataset"]
        family = row["coverage_item"]
        mapped = [
            item["candidate_rule_id"]
            for item in review_queue
            if dataset in item["dataset_variable_scope"].get("datasets", ())
            and _family_matches_queue_item(family, item)
        ][:20]
        rows.append(
            {
                "dataset": dataset,
                "coverage_family": family,
                "coverage_status": "INCOMPLETE",
                "missing_or_pending_reason": row["unresolved_conflicts"],
                "review_queue_rule_ids": tuple(mapped),
            }
        )
    return rows


def _rule_datasets(rule: CompiledRule) -> tuple[str, ...]:
    return tuple(dataset for dataset in V1_DATASETS if dataset in rule.scope.get("datasets", ()) or dataset in rule.expected_outputs)


def _rule_family(rule: CompiledRule) -> str:
    kind = rule.scope.get("candidate_kind")
    text = " ".join(str(value) for value in (rule.required_behavior, rule.evidence_locator.get("source_excerpt", ""))).lower()
    if kind == "dataset_structure" or "one record per" in text:
        return "structure_grain_keys"
    if kind == "controlled_terminology":
        return "controlled_terminology"
    if "traceability" in text:
        return "traceability"
    if "missing" in text or "imput" in text:
        return "missing_value_constraints"
    if "conformance" in text or rule.standard_id == "adam-conformance-rules":
        return "validation_conformance_requirements"
    if "treatment" in text or "emergent" in text or "baseline" in text or "aval" in text or "param" in text or "cnsr" in text:
        return "derivation_constraints"
    return "required_or_conditional_variables"


def _review_priority(rule: CompiledRule) -> int:
    score = 0
    if rule.classification == "STANDARD_REQUIRED":
        score += 50
    elif rule.classification == "STANDARD_GUIDED":
        score += 35
    if rule.standard_id in {"adam-model", "adamig", "adam-occds", "adam-bds-tte", "adam-conformance-rules"}:
        score += 20
    if _rule_family(rule) in {
        "structure_grain_keys",
        "required_or_conditional_variables",
        "derivation_constraints",
        "missing_value_constraints",
        "traceability",
        "validation_conformance_requirements",
    }:
        score += 15
    score += len(_rule_datasets(rule)) * 5
    return score


def _source_role(rule: CompiledRule) -> str:
    if rule.classification == "NON_NORMATIVE":
        return "validation_or_upstream_reference"
    return "primary_standard"


def _locator_without_long_fields(locator: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value for key, value in locator.items()
        if key not in {"source_excerpt", "duplicate_locators"}
    }


def _conflict_assessment(scope: str, related: list[CompiledRule]) -> str:
    if len(related) < 2:
        return "UNRESOLVED"
    excerpts = {_normalized_text(str(rule.required_behavior or rule.evidence_locator.get("source_excerpt", ""))) for rule in related}
    standards = {rule.standard_id for rule in related}
    if len(excerpts) == 1:
        return "DUPLICATE"
    if len(standards) == 1:
        return "DIFFERENT_SCOPE"
    if any(scope == dataset for dataset in V1_DATASETS):
        return "COMPLEMENTARY"
    return "UNRESOLVED"


def _family_matches_queue_item(family: str, item: dict[str, Any]) -> bool:
    text = " ".join(
        str(value)
        for value in (
            item.get("proposed_structured_requirement", {}),
            item.get("dataset_variable_scope", {}),
            item.get("short_source_excerpt", ""),
        )
    ).lower()
    if family in {"dataset_structure", "grain", "keys"}:
        return "structure" in text or "record" in text or "key" in text
    if family == "validation_conformance_requirements":
        return "conformance" in text or "validation" in text
    if family == "missing_value_constraints":
        return "missing" in text or "imput" in text
    if family == "source_traceability":
        return "traceability" in text or "source" in text
    return True


def _supporting_rules(
    rules: tuple[CompiledRule, ...],
    dataset: str,
    kinds: tuple[str, ...],
) -> tuple[CompiledRule, ...]:
    return tuple(
        rule
        for rule in rules
        if rule.extraction_status == "CANDIDATE_RULE"
        and rule.classification in {"STANDARD_REQUIRED", "STANDARD_GUIDED"}
        and rule.classification_status == "RESOLVED"
        and (dataset in rule.scope.get("datasets", ()) or dataset in rule.expected_outputs)
        and rule.scope.get("candidate_kind") in kinds
    )


def _citation_status(rule: CompiledRule) -> str:
    if not rule.official_url or not rule.local_relative_source_path or not rule.source_hash:
        return "CITATION_UNRESOLVED"
    locator = rule.evidence_locator
    if any(locator.get(key) for key in ("page", "section", "table", "row", "sheet")):
        return "RESOLVED"
    return "CITATION_UNRESOLVED"


def _citation_provenance_complete(rule: CompiledRule) -> bool:
    local_path = rule.local_relative_source_path
    return bool(
        rule.standard_id
        and rule.standard_name
        and rule.standard_version
        and local_path
        and not Path(str(local_path)).is_absolute()
        and rule.official_url
        and rule.source_hash
        and _citation_status(rule) == "RESOLVED"
    )


def _source_binding_status(locator: dict[str, Any], manifest: StandardManifest) -> str:
    excerpt = str(locator.get("source_excerpt") or "")
    if not excerpt:
        return "SOURCE_UNBOUND"
    if not any(locator.get(key) for key in ("page", "section", "table", "row", "sheet")):
        return "SOURCE_UNBOUND"
    if not manifest.official_url or not manifest.sha256:
        return "SOURCE_UNBOUND"
    return "SOURCE_BOUND"


def _reconstructed_rules(rules: tuple[CompiledRule, ...]) -> list[CompiledRule]:
    batch_findings = {candidate_id: decision for _, candidate_id, decision, _ in BATCH_001_ADJUDICATION_FINDINGS}
    reconstructed: list[CompiledRule] = []
    for rule in rules:
        if rule.extraction_status != "CANDIDATE_RULE":
            continue
        decision = batch_findings.get(rule.rule_id)
        if decision in {"REJECT", "EXCLUDE_FROM_V1_CORE", "NON_NORMATIVE_REFERENCE_ONLY"}:
            continue
        if decision == "REJECT_CURRENT_AND_SPLIT_INTO_ATOMIC_NAMING_RULES":
            reconstructed.extend(_split_atomic_naming_rules(rule))
            continue
        if decision == "RECONSTRUCT":
            reconstructed.append(_batch_001_reconstruction(rule))
            continue
        if rule.source_hash is None:
            continue
        if rule.source_section_classification not in NORMATIVE_SOURCE_SECTIONS and rule.rule_type != "CONFORMANCE_RULE":
            continue
        if rule.classification is None:
            continue
        reconstructed.append(rule)
    return reconstructed


def _batch_001_reconstruction(rule: CompiledRule) -> CompiledRule:
    if rule.rule_id == "adam-model:9b552da37cd3119e":
        return replace(
            rule,
            rule_id=f"{rule.rule_id}:reconstructed-adsl-one-record-per-subject",
            normalized_atomic_requirement="ADSL contains one record per subject and subject-level analysis attributes.",
            required_behavior="ADSL contains one record per subject and subject-level analysis attributes.",
            scope=_scope("", "dataset_structure", dataset_scope=("ADSL",), variable_scope=(), structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",)),
            dataset_scope=("ADSL",),
            structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",),
            variable_scope=(),
            grain="one record per subject",
            keys=("USUBJID",),
            semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING",
        )
    if rule.rule_id == "adam-conformance-rules:sheet-Rules Catalogue_row-728":
        return replace(
            rule,
            rule_id=f"{rule.rule_id}:reconstructed-adsl-usubjid-required",
            normalized_atomic_requirement="ADSL must contain USUBJID values required for subject identification.",
            required_behavior="ADSL must contain USUBJID values required for subject identification.",
            scope=_scope("", "conformance_rule", dataset_scope=("ADSL",), variable_scope=("USUBJID",), structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",)),
            dataset_scope=("ADSL",),
            structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",),
            variable_scope=("USUBJID",),
            keys=("USUBJID",),
            semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING",
        )
    if rule.rule_id == "adam-conformance-rules:sheet-Rules Catalogue_row-892":
        return replace(
            rule,
            rule_id=f"{rule.rule_id}:reconstructed-adsl-name-label",
            normalized_atomic_requirement='A dataset named ADSL must have the dataset label "Subject-Level Analysis Dataset".',
            required_behavior='A dataset named ADSL must have the dataset label "Subject-Level Analysis Dataset".',
            scope=_scope("", "conformance_rule", dataset_scope=("ADSL",), variable_scope=(), structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",)),
            dataset_scope=("ADSL",),
            structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET", "dataset metadata"),
            variable_scope=(),
            semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING",
        )
    if rule.rule_id == "adam-conformance-rules:sheet-Rules Catalogue_row-896":
        return replace(
            rule,
            rule_id=f"{rule.rule_id}:reconstructed-adsl-label-name",
            normalized_atomic_requirement='A dataset with label "Subject-Level Analysis Dataset" must be named ADSL.',
            required_behavior='A dataset with label "Subject-Level Analysis Dataset" must be named ADSL.',
            scope=_scope("", "conformance_rule", dataset_scope=("ADSL",), variable_scope=(), structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",)),
            dataset_scope=("ADSL",),
            structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET", "dataset metadata"),
            variable_scope=(),
            semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING",
        )
    return replace(rule, semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING")


def _split_atomic_naming_rules(rule: CompiledRule) -> list[CompiledRule]:
    return [
        replace(
            rule,
            rule_id=f"{rule.rule_id}:split-dataset-naming",
            normalized_atomic_requirement="ADaM dataset names follow standard dataset naming requirements when the cited standard states such a naming requirement.",
            required_behavior="ADaM dataset names follow standard dataset naming requirements when the cited standard states such a naming requirement.",
            scope=_scope("", "dataset_metadata", dataset_scope=rule.dataset_scope or _rule_datasets(rule), variable_scope=(), structure_scope=("dataset metadata",)),
            variable_scope=(),
            structure_scope=("dataset metadata",),
            semantic_reconstruction_status="SPLIT_FROM_BATCH_001_FINDING",
        ),
        replace(
            rule,
            rule_id=f"{rule.rule_id}:split-variable-naming",
            normalized_atomic_requirement="ADaM variable names follow standard variable naming requirements when the cited standard states such a naming requirement.",
            required_behavior="ADaM variable names follow standard variable naming requirements when the cited standard states such a naming requirement.",
            scope=_scope("", "variable_metadata", dataset_scope=rule.dataset_scope or _rule_datasets(rule), variable_scope=rule.variable_scope, structure_scope=("variable metadata",)),
            structure_scope=("variable metadata",),
            semantic_reconstruction_status="SPLIT_FROM_BATCH_001_FINDING",
        ),
    ]


def _apply_batch_001_artifact_reconstructions(
    root: Path,
    pack: CompiledKnowledgePack,
    reconstructed_rules: list[CompiledRule],
) -> list[CompiledRule]:
    batch_path = root / "review_queue" / "adjudication_batches_v2" / "batch_001.json"
    if not batch_path.exists():
        return reconstructed_rules
    existing_parent_ids = {rule.rule_id.split(":reconstructed-", 1)[0].split(":split-", 1)[0] for rule in reconstructed_rules}
    fingerprints = {
        item["standard_id"]: item
        for item in pack.manifest.source_manifest_fingerprints
        if item.get("standard_id")
    }
    batch = json.loads(batch_path.read_text(encoding="utf-8"))
    by_candidate = {item.get("candidate_rule_id"): item for item in batch.get("items", ())}
    additions: list[CompiledRule] = []
    for _, candidate_id, decision, instruction in BATCH_001_ADJUDICATION_FINDINGS:
        if decision not in {"RECONSTRUCT", "REJECT_CURRENT_AND_SPLIT_INTO_ATOMIC_NAMING_RULES"}:
            continue
        if candidate_id in existing_parent_ids:
            continue
        item = by_candidate.get(candidate_id)
        if not item:
            continue
        additions.extend(_rules_from_batch_001_item(item, fingerprints, decision, instruction))
    return sorted(
        [*reconstructed_rules, *additions],
        key=lambda rule: (rule.standard_id, rule.rule_id),
    )


def _rules_from_batch_001_item(
    item: dict[str, Any],
    fingerprints: dict[str, dict[str, Any]],
    decision: str,
    instruction: str | None,
) -> list[CompiledRule]:
    candidate_id = str(item["candidate_rule_id"])
    standard_id = candidate_id.split(":", 1)[0]
    fingerprint = fingerprints.get(standard_id, {})
    source_hash = fingerprint.get("sha256")
    local_path = fingerprint.get("local_path")
    if local_path and local_path.startswith("../../"):
        local_path = local_path.removeprefix("../../")
    exact_excerpt = str(item.get("short_source_excerpt") or "")
    locator = dict(item.get("locator") or {})
    locator["source_excerpt"] = _short_excerpt(exact_excerpt)
    locator["source_binding_status"] = "SOURCE_BOUND" if source_hash and any(locator.get(key) for key in ("page", "section", "table", "row", "sheet")) else "SOURCE_UNBOUND"
    locator["parser_compiler_version"] = KNOWLEDGE_COMPILER_VERSION
    if decision == "RECONSTRUCT":
        if candidate_id == "adam-model:9b552da37cd3119e":
            return [
                _batch_rule_from_item(
                    item,
                    fingerprint,
                    locator,
                    source_hash,
                    local_path,
                    suffix="reconstructed-adsl-one-record-per-subject",
                    normalized_requirement="ADSL contains one record per subject and subject-level analysis attributes.",
                    required_behavior="ADSL contains one record per subject and subject-level analysis attributes.",
                    dataset_scope=("ADSL",),
                    structure_scope=("SUBJECT_LEVEL_ANALYSIS_DATASET",),
                    variable_scope=(),
                    grain="one record per subject",
                    keys=("USUBJID",),
                    exact_excerpt=exact_excerpt,
                )
            ]
        return []
    if decision == "REJECT_CURRENT_AND_SPLIT_INTO_ATOMIC_NAMING_RULES":
        return [
            _batch_rule_from_item(
                item,
                fingerprint,
                locator,
                source_hash,
                local_path,
                suffix="split-dataset-naming",
                normalized_requirement="ADaM dataset names follow standard dataset naming requirements when the cited standard states such a naming requirement.",
                required_behavior="ADaM dataset names follow standard dataset naming requirements when the cited standard states such a naming requirement.",
                dataset_scope=tuple(item.get("datasets") or ()),
                structure_scope=("dataset metadata",),
                variable_scope=(),
                grain=None,
                keys=(),
                exact_excerpt=exact_excerpt,
            ),
            _batch_rule_from_item(
                item,
                fingerprint,
                locator,
                source_hash,
                local_path,
                suffix="split-variable-naming",
                normalized_requirement="ADaM variable names follow standard variable naming requirements when the cited standard states such a naming requirement.",
                required_behavior="ADaM variable names follow standard variable naming requirements when the cited standard states such a naming requirement.",
                dataset_scope=tuple(item.get("datasets") or ()),
                structure_scope=("variable metadata",),
                variable_scope=tuple(
                    value
                    for value in (item.get("scope") or {}).get("variables", ())
                    if value not in V1_DATASETS
                ),
                grain=None,
                keys=(),
                exact_excerpt=exact_excerpt,
            ),
        ]
    return []


def _batch_rule_from_item(
    item: dict[str, Any],
    fingerprint: dict[str, Any],
    locator: dict[str, Any],
    source_hash: str | None,
    local_path: str | None,
    *,
    suffix: str,
    normalized_requirement: str,
    required_behavior: str,
    dataset_scope: tuple[str, ...],
    structure_scope: tuple[str, ...],
    variable_scope: tuple[str, ...],
    grain: str | None,
    keys: tuple[str, ...],
    exact_excerpt: str,
) -> CompiledRule:
    candidate_id = str(item["candidate_rule_id"])
    return CompiledRule(
        rule_id=f"{candidate_id}:{suffix}",
        standard_id=candidate_id.split(":", 1)[0],
        standard_name=str(item.get("standard") or fingerprint.get("title") or ""),
        standard_version=item.get("standard_version") or fingerprint.get("version"),
        classification=item.get("classification") or item.get("proposed_classification") or "STANDARD_REQUIRED",
        scope=_scope("", str(item.get("rule_family") or "operative_rule"), dataset_scope=dataset_scope, variable_scope=variable_scope, structure_scope=structure_scope),
        applicability_conditions=tuple(),
        required_behavior=required_behavior,
        permitted_variations=tuple(),
        prohibited_behavior=tuple(),
        required_inputs=tuple(),
        expected_outputs=tuple(sorted(set(dataset_scope) | set(f"{dataset}.{variable}" for dataset in dataset_scope for variable in variable_scope))),
        grain=grain,
        keys=keys,
        validation_requirements=tuple(),
        evidence_locator=locator,
        local_relative_source_path=local_path,
        official_url=item.get("official_url") or fingerprint.get("official_url"),
        source_hash=source_hash,
        extraction_status="CANDIDATE_RULE",
        review_status="CANDIDATE",
        classification_status="RESOLVED" if source_hash else "UNRESOLVED",
        normative_strength="REQUIRED",
        source_section_classification="OPERATIVE_RULE",
        exact_source_excerpt=exact_excerpt,
        normalized_atomic_requirement=normalized_requirement,
        rule_type="OPERATIVE_RULE",
        dataset_scope=dataset_scope,
        structure_scope=structure_scope,
        variable_scope=variable_scope,
        validation_requirement=None,
        semantic_reconstruction_status="RECONSTRUCTED_FROM_BATCH_001_FINDING",
        excluded_reason=None,
        conformance_metadata=None,
    )


def _semantic_exclusions(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    batch_findings = {candidate_id: decision for _, candidate_id, decision, _ in BATCH_001_ADJUDICATION_FINDINGS}
    rows = []
    for rule in rules:
        reason = None
        decision = batch_findings.get(rule.rule_id)
        if decision in {"REJECT", "EXCLUDE_FROM_V1_CORE", "NON_NORMATIVE_REFERENCE_ONLY", "REJECT_CURRENT_AND_SPLIT_INTO_ATOMIC_NAMING_RULES"}:
            reason = decision
        elif rule.source_hash is None:
            reason = "source_hash_null"
        elif rule.source_section_classification in NON_OPERATIVE_SOURCE_SECTIONS:
            reason = f"non_operative_source_section:{rule.source_section_classification}"
        elif rule.classification is None:
            reason = "classification_unresolved"
        if reason:
            rows.append(
                {
                    "rule_id": rule.rule_id,
                    "standard_id": rule.standard_id,
                    "source_section_classification": rule.source_section_classification,
                    "reason": reason,
                    "review_status": rule.review_status,
                    "promotion_authorized": False,
                }
            )
    return rows


def _m3c_r1_reconciliation(
    source_rules: tuple[CompiledRule, ...],
    base_reconstructed: list[CompiledRule],
    reconstructed: list[CompiledRule],
    excluded: list[dict[str, Any]],
    citation_unresolved: list[CompiledRule],
    normalization_audit: dict[str, Any],
    *,
    schema_version: str = M3C_R1_SCHEMA_VERSION,
) -> dict[str, Any]:
    source_ids = {rule.rule_id for rule in source_rules}
    derived = [
        rule
        for rule in reconstructed
        if rule.semantic_reconstruction_status
        in {"RECONSTRUCTED_FROM_BATCH_001_FINDING", "SPLIT_FROM_BATCH_001_FINDING"}
    ]
    replacement_parent_ids = {
        parent_id
        for rule in derived
        if (parent_id := _reconstruction_parent_id(rule.rule_id)) in source_ids
    }
    artifact_additions = [
        rule
        for rule in derived
        if _reconstruction_parent_id(rule.rule_id) not in source_ids
    ]
    semantic_unaccounted = len(source_rules) - len(excluded) - len(base_reconstructed)
    citation_ready_count = len(reconstructed) - len(citation_unresolved)
    difference_key = f"difference_{len(source_rules)}_to_{len(reconstructed)}"
    return {
        "schema_version": schema_version,
        "source_total": len(source_rules),
        "source_total_definition": "Compiler rules entering semantic reconstruction after source-section filtering and source-level normalization.",
        "upstream_normalization": {
            "raw_candidate_count": normalization_audit.get("raw_candidate_count", 0),
            "removed_fragment_count": normalization_audit.get("removed_fragment_count", 0),
            "duplicate_records_detected": normalization_audit.get("merged_duplicate_count", 0),
            "duplicate_records_merged": normalization_audit.get("merged_duplicate_count", 0),
            "retained_candidate_count": normalization_audit.get("retained_candidate_count", 0),
            "note": f"The {len(source_rules)} semantic source total is already downstream of these source-level operations.",
        },
        "classifier_correction": {
            "publication_month_may_treated_as_normative_cue": False,
            "lowercase_may_remains_normative_cue": True,
        },
        "semantic_reconstruction_stages": [
            {"stage": "source_total", "removed": 0, "remaining": len(source_rules)},
            {"stage": "excluded", "removed": len(excluded), "remaining": len(source_rules) - len(excluded)},
            {"stage": "deduplicated", "removed": 0, "remaining": len(source_rules) - len(excluded)},
            {"stage": "merged", "removed": 0, "remaining": len(base_reconstructed)},
            {
                "stage": "batch_001_reconstruction",
                "count_neutral_parent_replacements": len(replacement_parent_ids),
                "added_reconstructed_rules_without_current_parent": len(artifact_additions),
                "remaining": len(reconstructed),
            },
            {
                "stage": "citation_blocked",
                "quarantined": len(citation_unresolved),
                "citation_resolved_candidates": citation_ready_count,
                "remaining_in_candidate_pack_including_quarantine": len(reconstructed),
            },
            {
                "stage": "reconstructed_candidate_pack",
                "candidate_count": len(reconstructed),
                "citation_resolved_count": citation_ready_count,
                "citation_unresolved_count": len(citation_unresolved),
                "runtime_pack_admitted_count": 0,
            },
        ],
        "excluded_by_reason": dict(sorted(Counter(item["reason"] for item in excluded).items())),
        "semantic_unaccounted_count": semantic_unaccounted,
        difference_key: {
            "source_total": len(source_rules),
            "reconstructed_candidate_count": len(reconstructed),
            "net_difference": len(source_rules) - len(reconstructed),
            "excluded": len(excluded),
            "count_neutral_replacements": len(replacement_parent_ids),
            "net_reconstruction_additions": len(artifact_additions),
            "explanation": "The net difference is excluded candidates minus reconstructed additions whose parent is absent from the current source pool; one-for-one replacements do not change the count.",
        },
        "review_status": "CANDIDATE",
        "promotion_authorized": False,
    }


def _reconstruction_parent_id(rule_id: str) -> str:
    for marker in (":reconstructed-", ":split-"):
        if marker in rule_id:
            return rule_id.split(marker, 1)[0]
    return rule_id


def _citation_unresolved_record(rule: CompiledRule) -> dict[str, Any]:
    missing = []
    if not rule.standard_version:
        missing.append("standard_version")
    if not rule.local_relative_source_path or Path(str(rule.local_relative_source_path)).is_absolute():
        missing.append("portable_local_relative_path")
    if not rule.official_url:
        missing.append("official_url")
    if not rule.source_hash:
        missing.append("source_sha256")
    if _citation_status(rule) != "RESOLVED":
        missing.append("resolvable_locator")
    return {
        "candidate_id": rule.rule_id,
        "standard_id": rule.standard_id,
        "standard_title": rule.standard_name,
        "standard_version": rule.standard_version,
        "source_section_classification": rule.source_section_classification,
        "citation_status": "CITATION_UNRESOLVED",
        "missing_provenance_fields": tuple(dict.fromkeys(missing)),
        "exact_source_excerpt": _exact_prefix(rule.exact_source_excerpt or ""),
        "locator": _locator_without_long_fields(rule.evidence_locator),
        "local_relative_path": rule.local_relative_source_path,
        "official_url": rule.official_url,
        "source_sha256": rule.source_hash,
        "review_queue_eligible": False,
        "runtime_pack_eligible": False,
        "review_status": "CANDIDATE",
        "promotion_authorized": False,
    }


def _source_section_spot_checks(
    rules: tuple[CompiledRule, ...],
    normalization_audit: dict[str, Any],
    *,
    sample_size: int,
    schema_version: str = M3C_R1_SCHEMA_VERSION,
) -> dict[str, Any]:
    operative = sorted(
        (
            rule
            for rule in rules
            if rule.extraction_status == "CANDIDATE_RULE"
            and rule.source_section_classification == "OPERATIVE_RULE"
        ),
        key=lambda rule: (rule.standard_id, rule.rule_id),
    )[:sample_size]
    front_matter = sorted(
        (
            item
            for item in normalization_audit.get("removed_fragments", ())
            if item.get("reason") == "non_operative_source_section:FRONT_MATTER"
        ),
        key=lambda item: (
            str(item.get("standard_id") or ""),
            _locator_key(item.get("locator") or {}),
            str(item.get("excerpt") or ""),
        ),
    )[:sample_size]
    return {
        "schema_version": schema_version,
        "selection_method": "First records after deterministic standard ID, candidate ID or locator ordering; classifications were not changed during sampling.",
        "OPERATIVE_RULE": [
            {
                "candidate_id": rule.rule_id,
                "source_type": "OPERATIVE_RULE",
                "exact_source_excerpt": _operative_spot_excerpt(rule),
                "locator": _locator_without_long_fields(rule.evidence_locator),
                "classification_reason": _operative_classification_reason(rule),
                "classification_changed_by_audit": False,
            }
            for rule in operative
        ],
        "FRONT_MATTER": [
            {
                "candidate_id": _front_matter_audit_id(item),
                "source_type": "FRONT_MATTER",
                "exact_source_excerpt": _exact_prefix(str(item.get("excerpt") or "")),
                "locator": item.get("locator") or {},
                "classification_reason": _front_matter_classification_reason(item),
                "classification_changed_by_audit": False,
            }
            for item in front_matter
        ],
        "counts": {
            "OPERATIVE_RULE": len(operative),
            "FRONT_MATTER": len(front_matter),
        },
    }


def _operative_classification_reason(rule: CompiledRule) -> str:
    locator = rule.evidence_locator
    if rule.rule_type == "CONFORMANCE_RULE" and str(locator.get("sheet") or "") == "Rules Catalogue":
        return "Existing classifier recognized an authoritative ADaM Conformance Rules catalogue row."
    text = rule.exact_source_excerpt or ""
    cue = _first_normative_cue(text)
    if cue:
        return f"Existing classifier retained the source block as operative because it contains the normative cue '{cue}'."
    return "Existing extractor/classifier supplied OPERATIVE_RULE; this audit preserved that classification unchanged."


def _operative_spot_excerpt(rule: CompiledRule) -> str:
    text = str(rule.exact_source_excerpt or rule.evidence_locator.get("source_excerpt") or "")
    for token in NORMATIVE_CUE_TOKENS:
        index = _normative_cue_index(text, token)
        if index >= 0:
            start = max(0, index - 180)
            return text[start : start + 480]
    return _exact_prefix(text)


def _front_matter_classification_reason(item: dict[str, Any]) -> str:
    locator = item.get("locator") or {}
    sheet = str(locator.get("sheet") or "")
    if sheet and sheet != "Rules Catalogue":
        return f"Existing workbook extractor explicitly classified the non-rules sheet '{sheet}' as FRONT_MATTER."
    text = str(item.get("excerpt") or "").lower()
    cue = next((token for token in ("copyright", "table of contents", "introduction / purpose") if token in text), None)
    if cue:
        return f"Existing classifier matched the front-matter cue '{cue}'."
    return "Existing extractor/classifier supplied FRONT_MATTER; this audit preserved that classification unchanged."


def _front_matter_audit_id(item: dict[str, Any]) -> str:
    standard_id = str(item.get("standard_id") or "source")
    identity = json.dumps(
        {"locator": item.get("locator") or {}, "excerpt": item.get("excerpt") or ""},
        sort_keys=True,
    )
    return f"{standard_id}:front-matter-audit-{_hash_text(identity)}"


def _exact_prefix(text: str, *, limit: int = 480) -> str:
    return text[:limit]


def _adsl_structure_grain_key_candidates(
    rules: tuple[CompiledRule, ...],
    *,
    limit: int,
) -> list[dict[str, Any]]:
    eligible = [
        rule
        for rule in rules
        if "ADSL" in rule.dataset_scope
        and _rule_family(rule) == "structure_grain_keys"
        and rule.classification in {"STANDARD_REQUIRED", "STANDARD_GUIDED"}
        and rule.classification_status == "RESOLVED"
        and rule.source_section_classification in NORMATIVE_SOURCE_SECTIONS
        and _citation_provenance_complete(rule)
        and rule.normalized_atomic_requirement
        and len(rule.normalized_atomic_requirement) <= 240
        and "operator" not in rule.scope
    ]
    eligible.sort(key=_adsl_core_selection_priority)
    selected: list[CompiledRule] = []
    seen_families = set()
    for rule in eligible:
        official_id = (rule.conformance_metadata or {}).get("official_rule_identifier")
        family_id = f"conformance:{official_id}" if official_id else rule.rule_id
        if family_id in seen_families:
            continue
        seen_families.add(family_id)
        selected.append(rule)
        if len(selected) >= limit:
            break
    return [_adsl_core_candidate_record(rule) for rule in selected]


def _adsl_core_selection_priority(rule: CompiledRule) -> tuple[int, int, int, str]:
    text = str(rule.normalized_atomic_requirement or "").lower()
    return (
        0 if "one record per subject" in text else 1,
        0 if rule.semantic_reconstruction_status == "RECONSTRUCTED_FROM_BATCH_001_FINDING" else 1,
        0 if rule.standard_id == "adam-model" else 1,
        rule.rule_id,
    )


def _adsl_core_candidate_record(rule: CompiledRule) -> dict[str, Any]:
    return {
        "candidate_id": rule.rule_id,
        "exact_source_excerpt": _atomic_supporting_excerpt(rule),
        "normalized_atomic_requirement": rule.normalized_atomic_requirement,
        "normative_strength": rule.normative_strength,
        "applicability_conditions": rule.applicability_conditions,
        "dataset_scope": rule.dataset_scope,
        "structure_scope": rule.structure_scope,
        "variable_scope": rule.variable_scope,
        "citation_provenance": {
            "standard_id": rule.standard_id,
            "standard_title": rule.standard_name,
            "standard_version": rule.standard_version,
            "local_relative_path": rule.local_relative_source_path,
            "official_url": rule.official_url,
            "locator": _locator_without_long_fields(rule.evidence_locator),
            "source_sha256": rule.source_hash,
            "parser_compiler_version": rule.evidence_locator.get("parser_compiler_version"),
            "citation_status": "RESOLVED",
        },
        "source_sha256": rule.source_hash,
        "review_status": "CANDIDATE",
        "semantic_review_status": "PENDING",
        "promotion_authorized": False,
        "runtime_pack_admitted": False,
    }


def _atomic_supporting_excerpt(rule: CompiledRule) -> str:
    text = str(rule.exact_source_excerpt or rule.evidence_locator.get("source_excerpt") or "")
    if rule.rule_type == "CONFORMANCE_RULE":
        fields = []
        keep = (
            "Rule ID:",
            "Rule Version:",
            "Rule Set:",
            "Scope:",
            "Variable or Item:",
            "Condition:",
            "Error Message:",
            "Rule:",
            "Cited Standard:",
            "Cited Section:",
        )
        for field in text.split(" | "):
            if field.startswith(keep):
                fields.append(field)
        if fields:
            return " | ".join(fields)
    for sentence in re.split(r"(?<=[.!?])\s+", text):
        if "one record per subject" in sentence.lower():
            return sentence
    return _exact_prefix(text)


def _semantic_rule_payload(rule: CompiledRule) -> dict[str, Any]:
    payload = asdict(rule)
    payload.update(
        {
            "exact_source_excerpt": rule.exact_source_excerpt,
            "normalized_atomic_requirement": rule.normalized_atomic_requirement,
            "rule_type": rule.rule_type,
            "normative_strength": rule.normative_strength,
            "dataset_scope": rule.dataset_scope,
            "structure_scope": rule.structure_scope,
            "variable_scope": rule.variable_scope,
            "validation_requirement": rule.validation_requirement,
            "review_status": "CANDIDATE",
        }
    )
    return payload


def _semantic_reconstruction_summary(
    all_rules: tuple[CompiledRule, ...],
    reconstructed_rules: list[CompiledRule],
    excluded: list[dict[str, Any]],
    normalization_audit: dict[str, Any],
) -> dict[str, Any]:
    section_counts: dict[str, int] = {}
    for rule in all_rules:
        key = rule.source_section_classification or "UNRESOLVED"
        section_counts[key] = section_counts.get(key, 0) + 1
    removed_by_non_operative_type: dict[str, int] = {}
    for item in normalization_audit.get("removed_fragments", ()):
        reason = str(item.get("reason") or "")
        prefix = "non_operative_source_section:"
        if reason.startswith(prefix):
            key = reason.removeprefix(prefix)
            removed_by_non_operative_type[key] = removed_by_non_operative_type.get(key, 0) + 1
            section_counts[key] = section_counts.get(key, 0) + 1
    removed_by_exclusion_reason: dict[str, int] = {}
    for item in excluded:
        reason = str(item.get("reason") or "unknown")
        removed_by_exclusion_reason[reason] = removed_by_exclusion_reason.get(reason, 0) + 1
    invented_operator_removed = sum(1 for rule in reconstructed_rules if "operator" not in rule.scope)
    scope_corrections = {
        "dataset_names_removed_from_variable_scope": sum(
            1 for rule in reconstructed_rules for value in rule.variable_scope if value in V1_DATASETS
        )
        == 0,
        "split_scope_fields_present": sum(
            1 for rule in reconstructed_rules
            if rule.dataset_scope is not None and rule.structure_scope is not None and rule.variable_scope is not None
        ),
    }
    citation_complete = [
        rule for rule in reconstructed_rules
        if rule.standard_id
        and rule.standard_name
        and rule.standard_version
        and rule.local_relative_source_path
        and not Path(str(rule.local_relative_source_path)).is_absolute()
        and rule.official_url
        and rule.source_hash
        and _citation_status(rule) == "RESOLVED"
    ]
    return {
        "schema_version": SEMANTIC_RECONSTRUCTION_SCHEMA_VERSION,
        "source_section_classification_counts": section_counts,
        "candidates_removed_by_non_operative_source_type": removed_by_non_operative_type,
        "candidates_removed_by_exclusion_reason": removed_by_exclusion_reason,
        "retained_candidate_count": len(reconstructed_rules),
        "excluded_candidate_count": len(excluded),
        "atomic_rules_created": sum(
            1
            for rule in reconstructed_rules
            if rule.semantic_reconstruction_status in {"RECONSTRUCTED_FROM_BATCH_001_FINDING", "SPLIT_FROM_BATCH_001_FINDING"}
        ),
        "invented_operators_removed": invented_operator_removed,
        "scope_corrections": scope_corrections,
        "citation_provenance": {
            "complete": len(citation_complete),
            "incomplete": len(reconstructed_rules) - len(citation_complete),
            "source_hash_null_rejected": all(rule.source_hash is not None for rule in reconstructed_rules),
        },
        "conformance_column_mapping": _conformance_column_mapping_summary(),
        "review_status": "CANDIDATE",
        "promotion_authorized": False,
    }


def _conformance_column_mapping_summary() -> dict[str, str]:
    return {
        "Rule ID": "official_rule_identifier",
        "Rule ID Version": "rule_version",
        "Related Rule(s)": "related_rules",
        "Rule Set": "rule_set",
        "Class/Subclass/Domain/Variable or Item": "scope",
        "Condition (Failure/Success)": "condition",
        "Natural Language Rule (Failure/Success Criteria)": "error_message",
        "Rule (Failure/Success Criteria)": "rule_expression",
        "Implementation Guide (Cited document)": "referenced_standard",
        "Cited Section": "referenced_section",
        "Cited Item": "referenced_item",
        "Cited Guidance": "cited_guidance",
        "Release Notes": "release_notes",
    }


def _semantic_review_input_freeze(
    root: Path,
    reconstructed_root: Path,
    pack: CompiledKnowledgePack,
    shard_index: list[dict[str, Any]],
) -> dict[str, Any]:
    review_queue_path = root / "review_queue" / "reconstructed_v1_core_review_queue.json"
    pack_payload = json.dumps(shard_index, sort_keys=True).encode("utf-8")
    return {
        "review_schema_version": SEMANTIC_RECONSTRUCTION_SCHEMA_VERSION,
        "compiler_version": pack.manifest.compiler_version,
        "pack_version": pack.manifest.pack_version,
        "candidate_pack_hash": hashlib.sha256(pack_payload).hexdigest(),
        "review_queue_hash": hashlib.sha256(review_queue_path.read_bytes()).hexdigest()
        if review_queue_path.exists()
        else None,
        "reconstructed_pack_root": str(reconstructed_root).replace("\\", "/"),
        "frozen_at": _utc_now(),
        "review_generation_status": "SUPERSEDED_BY_CANDIDATE_RECONSTRUCTION",
    }


def _write_rule_shards(rule_catalog: Path, rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_standard: dict[str, list[dict[str, Any]]] = {}
    for rule in rules:
        by_standard.setdefault(str(rule["standard_id"]), []).append(rule)
    index = []
    for standard_id in sorted(by_standard):
        shard_rules = sorted(by_standard[standard_id], key=lambda item: item["rule_id"])
        path = rule_catalog / f"{standard_id}.json"
        payload = {"rules": shard_rules}
        text = json.dumps(payload, indent=2, sort_keys=True)
        path.write_text(text, encoding="utf-8")
        resolved = sum(1 for rule in shard_rules if rule.get("classification_status") == "RESOLVED")
        relative = path.as_posix()
        index.append(
            {
                "file_path": relative,
                "standard_id": standard_id,
                "standard_version": shard_rules[0].get("standard_version") if shard_rules else None,
                "candidate_count": len(shard_rules),
                "resolved_count": resolved,
                "unresolved_count": len(shard_rules) - resolved,
                "content_hash": hashlib.sha256(text.encode("utf-8")).hexdigest(),
            }
        )
    return index


def _review_input_freeze(
    root: Path,
    pack: CompiledKnowledgePack,
    shard_index: list[dict[str, Any]],
) -> dict[str, Any]:
    review_queue_path = root / "review_queue" / "v1_core_review_queue.json"
    manifest_hashes = []
    registry_root = Path(pack.manifest.source_registry_root or ".")
    for path in sorted(registry_root.glob("*.yaml")):
        manifest_hashes.append(
            {
                "path": str(path),
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            }
        )
    pack_payload = json.dumps(shard_index, sort_keys=True).encode("utf-8")
    return {
        "review_schema_version": "m3r-phase-3d-review-v1",
        "compiler_version": pack.manifest.compiler_version,
        "pack_version": pack.manifest.pack_version,
        "candidate_pack_hash": hashlib.sha256(pack_payload).hexdigest(),
        "review_queue_hash": hashlib.sha256(review_queue_path.read_bytes()).hexdigest()
        if review_queue_path.exists()
        else None,
        "standards_manifest_hashes": manifest_hashes,
        "frozen_at": _utc_now(),
        "extraction_change_policy": "Any extraction or review-queue change requires a new review version.",
    }


def _normalized_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().lower()


def _hash_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def _locator_key(locator: dict[str, Any]) -> str:
    parts = []
    for key in ("sheet", "row", "page", "section"):
        value = locator.get(key)
        if value is not None:
            parts.append(f"{key}-{value}")
    return "_".join(parts) if parts else "locator-unresolved"


def _is_fragment_or_header(normalized_text: str) -> bool:
    if len(normalized_text) < 24:
        return True
    header_tokens = (
        "variable name variable label type codelist controlled terms core cdisc notes",
        "revision history date version",
        "copyright",
    )
    return normalized_text in header_tokens


def _summary(
    pack: CompiledKnowledgePack,
    unavailable: list[dict[str, Any]],
    normalization_audit: dict[str, Any],
) -> dict[str, Any]:
    counts: dict[str, dict[str, int]] = {}
    for rule in pack.rules:
        counts.setdefault(rule.standard_id, {})
        key = rule.classification or "CLASSIFICATION_UNRESOLVED"
        counts[rule.standard_id][key] = counts[rule.standard_id].get(key, 0) + 1
    citations = _citation_index(pack.rules)
    return {
        "standards_successfully_parsed": tuple(
            sorted({rule.standard_id for rule in pack.rules if rule.extraction_status == "CANDIDATE_RULE"})
        ),
        "standards_unavailable_or_mismatched": unavailable,
        "candidate_rule_counts_by_standard_and_classification": counts,
        "normalization_audit": normalization_audit,
        "coverage_status_by_dataset": _coverage_status_by_dataset(_coverage_matrix(pack.rules)),
        "unresolved_rules_or_conflicts": _unresolved(pack.rules),
        "study_decisions_still_required": STUDY_DECISION_TOPICS,
        "citation_resolution_summary": {
            "resolved": sum(1 for citation in citations if citation["citation_status"] == "RESOLVED"),
            "unresolved": sum(1 for citation in citations if citation["citation_status"] != "RESOLVED"),
        },
    }


def _coverage_status_by_dataset(rows: list[dict[str, Any]]) -> dict[str, str]:
    statuses = {}
    for dataset in V1_DATASETS:
        dataset_rows = [row for row in rows if row["dataset"] == dataset]
        statuses[dataset] = "INCOMPLETE"
    return statuses


def _unresolved(rules: tuple[CompiledRule, ...]) -> list[dict[str, Any]]:
    rows = []
    for rule in rules:
        if (
            rule.extraction_status != "CANDIDATE_RULE"
            or rule.classification_status != "RESOLVED"
            or _citation_status(rule) != "RESOLVED"
        ):
            rows.append(
                {
                    "rule_id": rule.rule_id,
                    "standard_id": rule.standard_id,
                    "classification": rule.classification,
                    "extraction_status": rule.extraction_status,
                    "citation_status": _citation_status(rule),
                    "review_status": rule.review_status,
                }
            )
    return rows


def _fingerprint(manifest: StandardManifest, integrity: dict[str, Any], extracted_at: str) -> dict[str, Any]:
    return {
        "standard_id": manifest.id,
        "title": manifest.title,
        "version": manifest.version,
        "source_role": manifest.role,
        "sha256": manifest.sha256,
        "sha256_status": integrity.get("sha256_status"),
        "local_path": manifest.local_path,
        "official_url": manifest.official_url,
        "extraction_timestamp": extracted_at,
        "parser_compiler_version": KNOWLEDGE_COMPILER_VERSION,
        "identity_status": integrity["status"],
    }


def _problem(manifest: StandardManifest, integrity: dict[str, Any]) -> dict[str, Any]:
    return {
        "standard_id": manifest.id,
        "title": manifest.title,
        "version": manifest.version,
        "source_role": manifest.role,
        "local_path": manifest.local_path,
        "official_url": manifest.official_url,
        "status": integrity["status"],
        "reason": integrity.get("reason"),
    }


def _relative_source(root: Path, path: Path) -> str:
    resolved_root = root.resolve()
    resolved_path = path.resolve()
    for base in (resolved_root, resolved_root.parent, resolved_root.parent.parent):
        try:
            return resolved_path.relative_to(base).as_posix()
        except ValueError:
            continue
    return resolved_path.name


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _short_excerpt(value: object, limit: int = 420) -> str:
    text = re.sub(r"\s+", " ", str(value)).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()
